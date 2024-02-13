# Файл views.py
### Recomended version of Python is 3.12, because psycopg2-binary is stable on this or higheir version.
import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.backends.db import SessionStore
import datetime as dt
from io import BytesIO
from django.http import FileResponse
from django.db import transaction

from .database_requests_table1 import (getDataTableForAllTime,
                                       getUserInfoFromDB,
                                       getUserInfoFromDBDataset,
                                       getTranzitUserInfoFromDB,
                                       getAllTranzitUserInfoFromDB,
                                       getMaxWarehouseQty,
                                       getMaxWarehouseAllQty,
                                       getNormsWarehouseQty,
                                       getNormsWarehouseAllQty,
                                       getReidUserInfoFromDB,
                                       getReidAllInfoFromDB,
                                       getMaxWarehouseAllQtyNotST)

from .database_requests_table2 import (getContaunerUserInfoFromDB,
                                       getContaunerInfoFromDBAll)

from .database_requests_table3_4 import (getWagonsUserInfoFromDB,
                                         getWagonsInfoFromDBAll,
                                         getWagonsUserInfoFromDBFE,
                                         getWagonsInfoFromDBAllFE)

from .database_requests_table5 import (getTransportUserInfoFromDB,
                                       getTransportInfoFromDBAll,
                                       getTransportInfoFromDBNotSTAll)

from .models import (DailyMonitoringUserData,
                     ConstantUserData,
                     DailyMonitoringUserContainers,
                     DailyMonitoringUserWagons,
                     DailyMonitoringUserWagonsFE,
                     DailyMonitoringUserTransport)

from .utils import (AllQtyPercent,
                    TransportPercent,
                    AllQtyCalculator,
                    TransportCalculator,
                    ReidAllStr,
                    PortAllStr,
                    AutoCalculator,
                    is_stevedor,
                    NanCheck,
                    connection)

# 1 day in timedelta data format
DAYDELTA = dt.timedelta(days=1,
                           seconds=0,
                           microseconds=0,
                           milliseconds=0,
                           minutes=0,
                           hours=0,
                           weeks=0)


# home page
@login_required(login_url='')
def home(request):
    if request.user.id != 1:
        if is_stevedor(request.user):
            return render(request, 'home.html')
        else:
            return render(request, 'home_notstevedor.html')
    else:
        return render(request, 'admin.html')

# register page
@transaction.atomic
@login_required(login_url='')
def register(request):
    if request.user.id == 1:
        if request.method == 'POST':
            try:
                username = request.POST['username']
                password = request.POST['password']
                norms = request.POST['norms']
                max = request.POST['max']

                # Создание нового пользователя
                user = User.objects.create_user(username=username, password=password)
                user.save()
                ConstantUserData.objects.create(db_userid=user.id,
                                                db_norms=norms,
                                                db_max=max)
                return redirect('login')
            except:
                return render(request, 'register.html')
        return render(request, 'register.html')
    else:
        return redirect('login')
    ####

# login page
def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        # Проверка введенных данных
        user = authenticate(request,
                            username=username,
                            password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return redirect('login')
    return render(request, 'login.html')

# logout realization
def user_logout(request):
    # Выход пользователя
    logout(request)
    return redirect('home')

@login_required(login_url='')
def table1_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        #WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        ImportIn = params.get('ImportIn')
        ImportOut = params.get('ImportOut')
        ExportIn = params.get('ExportIn')
        ExportOut = params.get('ExportOut')
        TransitIn = params.get('TransitIn')
        TransitOut = params.get('TransitOut')
        ExportEmpty = params.get('ExportEmpty')
        OtherEmpty = params.get('OtherEmpty')
        UnloadReidlin = params.get('UnloadReidlin')
        UnloadReidtramp = params.get('UnloadReidtramp')
        LoadingReidLin = params.get('LoadingReidLin')
        LoadingReidTramp = params.get('LoadingReidTramp')
        UnloadPortlin = params.get('UnloadPortlin')
        UnloadPorttramp = params.get('UnloadPorttramp')
        LoadingPortLin = params.get('LoadingPortLin')
        LoadingPortTramp = params.get('LoadingPortTramp')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            ImportIn = [request.POST['ImportIn']]
            ImportOut = [request.POST['ImportOut']]
            ExportIn = [request.POST['ExportIn']]
            ExportOut = [request.POST['ExportOut']]
            TransitIn = [request.POST['TransitIn']]
            TransitOut = [request.POST['TransitOut']]
            ExportEmpty = [request.POST['ExportEmpty']]
            OtherEmpty = [request.POST['OtherEmpty']]
            UnloadReidlin = [request.POST['UnloadReidlin']]
            UnloadReidtramp = [request.POST['UnloadReidtramp']]
            LoadingReidLin = [request.POST['LoadingReidLin']]
            LoadingReidTramp = [request.POST['LoadingReidTramp']]
            UnloadPortlin = [request.POST['UnloadPortlin']]
            UnloadPorttramp = [request.POST['UnloadPorttramp']]
            LoadingPortLin = [request.POST['LoadingPortLin']]
            LoadingPortTramp = [request.POST['LoadingPortTramp']]
            request.session['parameters'] = {
                'date2' : date2,
                'ImportIn': ImportIn,
                'ImportOut': ImportOut,
                'ExportIn': ExportIn,
                'ExportOut': ExportOut,
                'TransitIn': TransitIn,
                'TransitOut': TransitOut,
                'ExportEmpty': ExportEmpty,
                'OtherEmpty': OtherEmpty,
                'UnloadReidlin': UnloadReidlin,
                'UnloadReidtramp': UnloadReidtramp,
                'LoadingReidLin': LoadingReidLin,
                'LoadingReidTramp': LoadingReidTramp,
                'UnloadPortlin' : UnloadPortlin,
                'UnloadPorttramp': UnloadPorttramp,
                'LoadingPortLin': LoadingPortLin,
                'LoadingPortTramp': LoadingPortTramp
            }
            redirect_url = f'/success_table1/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table1_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'ImportIn': ImportIn[0],
                                                        'ImportOut': ImportOut[0],
                                                        'ExportIn': ExportIn[0],
                                                        'ExportOut': ExportOut[0],
                                                        'TransitIn': TransitIn[0],
                                                        'TransitOut': TransitOut[0],
                                                        'ExportEmpty': ExportEmpty[0],
                                                        'OtherEmpty': OtherEmpty[0],
                                                        'UnloadReidlin': UnloadReidlin[0],
                                                        'UnloadReidtramp': UnloadReidtramp[0],
                                                        'LoadingReidLin': LoadingReidLin[0],
                                                        'LoadingReidTramp': LoadingReidTramp[0],
                                                        'UnloadPortlin' : UnloadPortlin[0],
                                                        'UnloadPorttramp': UnloadPorttramp[0],
                                                        'LoadingPortLin': LoadingPortLin[0],
                                                        'LoadingPortTramp': LoadingPortTramp[0]
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            ImportIn = [request.POST['ImportIn']]
            ImportOut = [request.POST['ImportOut']]
            ExportIn = [request.POST['ExportIn']]
            ExportOut = [request.POST['ExportOut']]
            TransitIn = [request.POST['TransitIn']]
            TransitOut = [request.POST['TransitOut']]
            ExportEmpty = [request.POST['ExportEmpty']]
            OtherEmpty = [request.POST['OtherEmpty']]
            UnloadReidlin = [request.POST['UnloadReidlin']]
            UnloadReidtramp = [request.POST['UnloadReidtramp']]
            LoadingReidLin = [request.POST['LoadingReidLin']]
            LoadingReidTramp = [request.POST['LoadingReidTramp']]
            UnloadPortlin = [request.POST['UnloadPortlin']]
            UnloadPorttramp = [request.POST['UnloadPorttramp']]
            LoadingPortLin = [request.POST['LoadingPortLin']]
            LoadingPortTramp = [request.POST['LoadingPortTramp']]
            request.session['parameters'] = {
                'date2': date2,
                'ImportIn': ImportIn,
                'ImportOut': ImportOut,
                'ExportIn': ExportIn,
                'ExportOut': ExportOut,
                'TransitIn': TransitIn,
                'TransitOut': TransitOut,
                'ExportEmpty': ExportEmpty,
                'OtherEmpty' : OtherEmpty,
                'UnloadReidlin': UnloadReidlin,
                'UnloadReidtramp': UnloadReidtramp,
                'LoadingReidLin': LoadingReidLin,
                'LoadingReidTramp': LoadingReidTramp,
                'UnloadPortlin' : UnloadPortlin,
                'UnloadPorttramp': UnloadPorttramp,
                'LoadingPortLin': LoadingPortLin,
                'LoadingPortTramp': LoadingPortTramp
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table1/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            conn = connection()
            Userdata = getUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table1_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ImportIn' : NanCheck(Userdata[0][0]),
                'ImportOut' : NanCheck(Userdata[0][1]),
                'ExportIn' : NanCheck(Userdata[0][2]),
                'ExportOut' : NanCheck(Userdata[0][3]),
                'TransitIn' : NanCheck(Userdata[0][4]),
                'TransitOut' : NanCheck(Userdata[0][5]),
                'ExportEmpty' : NanCheck(Userdata[0][6]),
                'OtherEmpty' : NanCheck(Userdata[0][7]),
                'UnloadReidlin' : NanCheck(Userdata[0][8]),
                'UnloadReidtramp' : NanCheck(Userdata[0][9]),
                'LoadingReidLin' : NanCheck(Userdata[0][10]),
                'LoadingReidTramp' : NanCheck(Userdata[0][11]),
                'UnloadPortlin' : NanCheck(Userdata[0][12]),
                'UnloadPorttramp' : NanCheck(Userdata[0][13]),
                'LoadingPortLin' : NanCheck(Userdata[0][13]),
                'LoadingPortTramp' : NanCheck(Userdata[0][14])
            })
        except:
            return render(request, 'table1_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ImportIn': 0,
                'ImportOut': 0,
                'ExportIn': 0,
                'ExportOut': 0,
                'TransitIn': 0,
                'TransitOut': 0,
                'ExportEmpty': 0,
                'OtherEmpty': 0,
                'UnloadReidlin': 0,
                'UnloadReidtramp': 0,
                'LoadingReidLin': 0,
                'LoadingReidTramp': 0,
                'UnloadPortlin': 0,
                'UnloadPorttramp': 0,
                'LoadingPortLin': 0,
                'LoadingPortTramp': 0
            })

@transaction.atomic
@login_required(login_url='')
def success_table1(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')

        ImportIn = params.get('ImportIn')
        ImportOut = params.get('ImportOut')
        ExportIn = params.get('ExportIn')
        ExportOut = params.get('ExportOut')
        TransitIn = params.get('TransitIn')
        TransitOut = params.get('TransitOut')
        ExportEmpty = params.get('ExportEmpty')
        OtherEmpty = params.get('OtherEmpty')
        UnloadReidlin = params.get('UnloadReidlin')
        UnloadReidtramp = params.get('UnloadReidtramp')
        LoadingReidLin = params.get('LoadingReidLin')
        LoadingReidTramp = params.get('LoadingReidTramp')
        UnloadPortlin = params.get('UnloadPortlin')
        UnloadPorttramp = params.get('UnloadPorttramp')
        LoadingPortLin = params.get('LoadingPortLin')
        LoadingPortTramp = params.get('LoadingPortTramp')
        ImportIn[0] = NanCheck(ImportIn[0])
        ImportOut[0] = NanCheck(ImportOut[0])
        ExportIn[0] = NanCheck(ExportIn[0])
        ExportOut[0] = NanCheck(ExportOut[0])
        TransitIn[0] = NanCheck(TransitIn[0])
        TransitOut[0] = NanCheck(TransitOut[0])
        ExportEmpty[0] = NanCheck(ExportEmpty[0])
        OtherEmpty[0] = NanCheck(OtherEmpty[0])
        UnloadReidlin[0] = NanCheck(UnloadReidlin[0])
        UnloadReidtramp[0] = NanCheck(UnloadReidtramp[0])
        LoadingReidLin[0] = NanCheck(LoadingReidLin[0])
        LoadingReidTramp[0] = NanCheck(LoadingReidTramp[0])
        UnloadPortlin[0] = NanCheck(UnloadPortlin[0])
        UnloadPorttramp[0] = NanCheck(UnloadPorttramp[0])
        LoadingPortLin[0] = NanCheck(LoadingPortLin[0])
        LoadingPortTramp[0] = NanCheck(LoadingPortTramp[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserData.objects.filter(
                date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                db_userid = request.user.id).update(
                db_importin=int(ImportIn[0]),
                db_importout=int(ImportOut[0]),
                db_exportin=int(ExportIn[0]),
                db_exportout=int(ExportOut[0]),
                db_transitin=int(TransitIn[0]),
                db_transitout=int(TransitOut[0]),
                db_exportempty=int(ExportEmpty[0]),
                db_otherempty=int(OtherEmpty[0]),
                db_unload_reid_lin=int(UnloadReidlin[0]),
                db_unload_reid_tramp=int(UnloadReidtramp[0]),
                db_loading_reid_lin=int(LoadingReidLin[0]),
                db_loading_reid_tramp=int(LoadingReidTramp[0]),
                db_loading_port_lin=int(UnloadPortlin[0]),
                db_loading_port_tramp=int(UnloadPorttramp[0]),
                db_unload_port_lin=int(LoadingPortLin[0]),
                db_unload_port_tramp=int(LoadingPortTramp[0])
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserData.objects.create(
                date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                          db_userid = request.user.id,
                                          db_importin= int(ImportIn[0]),
                                          db_importout = int(ImportOut[0]),
                                          db_exportin = int(ExportIn[0]),
                                          db_exportout = int(ExportOut[0]),
                                          db_transitin = int(TransitIn[0]),
                                          db_transitout = int(TransitOut[0]),
                                          db_exportempty = int(ExportEmpty[0]),
                                          db_otherempty = int(OtherEmpty[0]),
                                          db_unload_reid_lin=int(UnloadReidlin[0]),
                                          db_unload_reid_tramp=int(UnloadReidtramp[0]),
                                          db_loading_reid_lin=int(LoadingReidLin[0]),
                                          db_loading_reid_tramp=int(LoadingReidTramp[0]),
                                          db_loading_port_lin=int(UnloadPortlin[0]),
                                          db_loading_port_tramp=int(UnloadPorttramp[0]),
                                          db_unload_port_lin=int(LoadingPortLin[0]),
                                          db_unload_port_tramp=int(LoadingPortTramp[0])
                                          )

        return render(request, 'success_table1.html', {
                                                'date2': date2[0][0],
                                                'ImportIn': ImportIn[0],
                                                'ImportOut': ImportOut[0],
                                                'ExportIn': ExportIn[0],
                                                'ExportOut': ExportOut[0],
                                                'TransitIn': TransitIn[0],
                                                'TransitOut': TransitOut[0],
                                                'ExportEmpty': ExportEmpty[0],
                                                'OtherEmpty': OtherEmpty[0],
                                                'UnloadReidlin': UnloadReidlin[0],
                                                'UnloadReidtramp': UnloadReidtramp[0],
                                                'LoadingReidLin': LoadingReidLin[0],
                                                'LoadingReidTramp': LoadingReidTramp[0],
                                                'UnloadPortlin' : UnloadPortlin[0],
                                                'UnloadPorttramp': UnloadPorttramp[0],
                                                'LoadingPortLin': LoadingPortLin[0],
                                                'LoadingPortTramp': LoadingPortTramp[0],
                                                'user': request.user})


###File download realization
@login_required(login_url='')
def download(request):
    if request.user.id == 1:
        wb = openpyxl.load_workbook('./WS3_s.xlsx')
        ws = wb.get_sheet_by_name('Шаблон1')
        params = request.session.get('parameters', {})
        date = params.get('date1')
        try:
            conn = connection()
            #
            # TestUser1
            #
            # table 1
            Tranzit1 = getTranzitUserInfoFromDB(2, date,conn)
            User1data = getUserInfoFromDBDataset(2, date,conn)
            AllQty1User = AllQtyCalculator(User1data, Tranzit1)
            Reid_info1 = getReidUserInfoFromDB(2, date,conn)
            # table 2
            ContainerData1 = getContaunerUserInfoFromDB(2, date,conn)
            ContainerDataNow1 = getContaunerUserInfoFromDB(2,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData1 = getWagonsUserInfoFromDB(2, date,conn)
            # table4
            WagonsDataFE1 = getWagonsUserInfoFromDBFE(2, date,conn)
            # table5
            TransportUserInfo1 = getTransportUserInfoFromDB(2, date, conn)

            #
            # TestUser2
            #
            # table 1
            Tranzit2 = getTranzitUserInfoFromDB(3, date,conn)
            User2data = getUserInfoFromDBDataset(3, date,conn)
            AllQty2User = AllQtyCalculator(User2data, Tranzit2)
            Reid_info2 = getReidUserInfoFromDB(3, date,conn)

            # table 2
            ContainerData2 = getContaunerUserInfoFromDB(3, date,conn)
            ContainerDataNow2 = getContaunerUserInfoFromDB(3,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData2 = getWagonsUserInfoFromDB(3, date,conn)
            # table4
            WagonsDataFE2 = getWagonsUserInfoFromDBFE(3, date,conn)
            # table5
            TransportUserInfo2 = getTransportUserInfoFromDB(3, date, conn)


            #
            # 3 user
            #

            # table 1
            Tranzit3 = getTranzitUserInfoFromDB(4, date,conn)
            User3data = getUserInfoFromDBDataset(4, date,conn)
            AllQty3User = AllQtyCalculator(User3data, Tranzit3)
            Reid_info3 = getReidUserInfoFromDB(4, date,conn)

            # table 2
            ContainerData3 = getContaunerUserInfoFromDB(4, date,conn)
            ContainerDataNow3 = getContaunerUserInfoFromDB(4,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData3 = getWagonsUserInfoFromDB(4, date,conn)
            # table4
            WagonsDataFE3 = getWagonsUserInfoFromDBFE(4, date,conn)
            # table5
            TransportUserInfo3 = getTransportUserInfoFromDB(4, date, conn)


            #
            # 4 user
            #

            # table 1
            Tranzit4 = getTranzitUserInfoFromDB(5, date,conn)
            User4data = getUserInfoFromDBDataset(5, date,conn)
            AllQty4User = AllQtyCalculator(User4data, Tranzit4)

            Reid_info4 = getReidUserInfoFromDB(5, date,conn)

            # table 2
            ContainerData4 = getContaunerUserInfoFromDB(5, date,conn)
            ContainerDataNow4 = getContaunerUserInfoFromDB(5,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData4 = getWagonsUserInfoFromDB(5, date,conn)
            # table4
            WagonsDataFE4 = getWagonsUserInfoFromDBFE(5, date,conn)
            # table5
            TransportUserInfo4 = getTransportUserInfoFromDB(5, date, conn)


            #
            # 5 user
            #

            # table 1
            Tranzit5 = getTranzitUserInfoFromDB(6, date,conn)
            User5data = getUserInfoFromDBDataset(6, date,conn)
            AllQty5User = AllQtyCalculator(User5data, Tranzit5)

            Reid_info5 = getReidUserInfoFromDB(6, date,conn)

            # table 2
            ContainerData5 = getContaunerUserInfoFromDB(6, date,conn)
            ContainerDataNow5 = getContaunerUserInfoFromDB(6,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData5 = getWagonsUserInfoFromDB(6, date,conn)
            # table4
            WagonsDataFE5 = getWagonsUserInfoFromDBFE(6, date,conn)
            # table5
            TransportUserInfo5 = getTransportUserInfoFromDB(6, date,conn)

            #
            # 6 user
            #

            # table 1
            Tranzit6 = getTranzitUserInfoFromDB(7, date,conn)
            User6data = getUserInfoFromDBDataset(7, date,conn)
            AllQty6User = AllQtyCalculator(User6data, Tranzit6)
            Reid_info6 = getReidUserInfoFromDB(7, date,conn)

            # table 2
            ContainerData6 = getContaunerUserInfoFromDB(7, date,conn)
            ContainerDataNow6 = getContaunerUserInfoFromDB(7,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData6 = getWagonsUserInfoFromDB(7, date,conn)
            # table4
            WagonsDataFE6 = getWagonsUserInfoFromDBFE(7, date,conn)
            # table5
            TransportUserInfo6 = getTransportUserInfoFromDB(7, date,conn)


            #
            # 7 user
            #

            # table 1
            Tranzit7 = getTranzitUserInfoFromDB(8, date,conn)
            User7data = getUserInfoFromDBDataset(8, date,conn)
            AllQty7User = AllQtyCalculator(User7data, Tranzit7)

            Reid_info7 = getReidUserInfoFromDB(8, date,conn)

            # table 2
            ContainerData7 = getContaunerUserInfoFromDB(8, date,conn)
            ContainerDataNow7 = getContaunerUserInfoFromDB(8,
                                                           (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
                                                               '%Y-%m-%d'),conn)
            # table3
            WagonsData7 = getWagonsUserInfoFromDB(8, date,conn)
            # table4
            WagonsDataFE7 = getWagonsUserInfoFromDBFE(8, date,conn)
            # table5
            TransportUserInfo7 = getTransportUserInfoFromDB(8, date,conn)



            #####
            #######6 table##### 2 table in dataset
            #####

            # 8 user

            TransportUserInfo8 = getTransportUserInfoFromDB(9, date,conn)


            # 9 user

            TransportUserInfo9 = getTransportUserInfoFromDB(10, date,conn)


            # 10 user

            TransportUserInfo10 = getTransportUserInfoFromDB(11, date,conn)


            # 11 user

            TransportUserInfo11 = getTransportUserInfoFromDB(12, date,conn)


            # 12 user

            TransportUserInfo12 = getTransportUserInfoFromDB(13, date,conn)


            # 13 user

            TransportUserInfo13 = getTransportUserInfoFromDB(14, date,conn)


            # 14 user

            TransportUserInfo14 = getTransportUserInfoFromDB(15, date,conn)

            # date

            ws['L4'] = (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime('%d.%m.%Y')

            for i in range(8,15):
                ws[f'C{i}'] = getMaxWarehouseQty(i-6,conn)[0][0]  #Полная проектная емкость складов (100%)
                ws[f'D{i}'] = getNormsWarehouseQty(i-6,conn)[0][0] #Емкость технологического накопления контейнеров (90%)

                #Факт, контейнеры

            ws['E8'] = AllQty1User
            ws['E9'] = AllQty2User
            ws['E10'] = AllQty3User
            ws['E11'] = AllQty4User
            ws['E12'] = AllQty5User
            ws['E13'] = AllQty6User
            ws['E14'] = AllQty7User

            #в т.ч растаможенных

            ws['G8'] = NanCheck(ContainerData1[0][0])
            ws['G9'] = NanCheck(ContainerData2[0][0])
            ws['G10'] = NanCheck(ContainerData3[0][0])
            ws['G11'] = NanCheck(ContainerData4[0][0])
            ws['G12'] = NanCheck(ContainerData5[0][0])
            ws['G13'] = NanCheck(ContainerData6[0][0])
            ws['G14'] = NanCheck(ContainerData7[0][0])

            #Факт прибытия всеми видами транспорта

            ws['I8'] = NanCheck(TransportUserInfo1[0][4]) + NanCheck(TransportUserInfo1[0][5]) + NanCheck(
                TransportUserInfo1[0][6]) + NanCheck(TransportUserInfo1[0][7])
            ws['I9'] = NanCheck(TransportUserInfo2[0][4]) + NanCheck(TransportUserInfo2[0][5]) + NanCheck(
                TransportUserInfo2[0][6]) + NanCheck(TransportUserInfo2[0][7])
            ws['I10'] = NanCheck(TransportUserInfo3[0][4]) + NanCheck(TransportUserInfo3[0][5]) + NanCheck(
                TransportUserInfo3[0][6]) + NanCheck(TransportUserInfo3[0][7])
            ws['I11'] = NanCheck(TransportUserInfo4[0][4]) + NanCheck(TransportUserInfo4[0][5]) + NanCheck(
                TransportUserInfo4[0][6]) + NanCheck(TransportUserInfo4[0][7])
            ws['I12'] = NanCheck(TransportUserInfo5[0][4]) + NanCheck(TransportUserInfo5[0][5]) + NanCheck(
                TransportUserInfo5[0][6]) + NanCheck(TransportUserInfo5[0][7])
            ws['I13'] = NanCheck(TransportUserInfo6[0][4]) + NanCheck(TransportUserInfo6[0][5]) + NanCheck(
                TransportUserInfo6[0][6]) + NanCheck(TransportUserInfo6[0][7])
            ws['I14'] = NanCheck(TransportUserInfo7[0][4]) + NanCheck(TransportUserInfo7[0][5]) + NanCheck(
                TransportUserInfo7[0][6]) + NanCheck(TransportUserInfo7[0][7])

            # Факт убытия всеми видами транспорта

            ws['L8'] = NanCheck(TransportUserInfo1[0][0]) + NanCheck(TransportUserInfo1[0][1]) + NanCheck(
                TransportUserInfo1[0][2]) + NanCheck(TransportUserInfo1[0][3])
            ws['L9'] = NanCheck(TransportUserInfo2[0][0]) + NanCheck(TransportUserInfo2[0][1]) + NanCheck(
                TransportUserInfo2[0][2]) + NanCheck(TransportUserInfo2[0][3])
            ws['L10'] = NanCheck(TransportUserInfo3[0][0]) + NanCheck(TransportUserInfo3[0][1]) + NanCheck(
                TransportUserInfo3[0][2]) + NanCheck(TransportUserInfo3[0][3])
            ws['L11'] = NanCheck(TransportUserInfo4[0][0]) + NanCheck(TransportUserInfo4[0][1]) + NanCheck(
                TransportUserInfo4[0][2]) + NanCheck(TransportUserInfo4[0][3])
            ws['L12'] = NanCheck(TransportUserInfo5[0][0]) + NanCheck(TransportUserInfo5[0][1]) + NanCheck(
                TransportUserInfo5[0][2]) + NanCheck(TransportUserInfo5[0][3])
            ws['L13'] = NanCheck(TransportUserInfo6[0][0]) + NanCheck(TransportUserInfo6[0][1]) + NanCheck(
                TransportUserInfo6[0][2]) + NanCheck(TransportUserInfo6[0][3])
            ws['L14'] = NanCheck(TransportUserInfo7[0][0]) + NanCheck(TransportUserInfo7[0][1]) + NanCheck(
                TransportUserInfo7[0][2]) + NanCheck(TransportUserInfo7[0][3])


            #####2 table
            for i in range(20,27):
                ws[f'C{i}'] = getMaxWarehouseQty(i-11,conn)[0][0]  #Максимальная емкость складов

            #Факт загрузки экспортно-импортных контейнеров на начало текущей даты
            ws['D20'] = NanCheck(TransportUserInfo8[0][8])
            ws['D21'] = NanCheck(TransportUserInfo9[0][8])
            ws['D22'] = NanCheck(TransportUserInfo10[0][8])
            ws['D23'] = NanCheck(TransportUserInfo11[0][8])
            ws['D24'] = NanCheck(TransportUserInfo12[0][8])
            ws['D25'] = NanCheck(TransportUserInfo13[0][8])
            ws['D26'] = NanCheck(TransportUserInfo14[0][8])

            #Всего убыло всеми видами транспорта	Всего прибыло всеми видами
            ws['F20'] = NanCheck(TransportUserInfo8[0][0]) + NanCheck(
                TransportUserInfo8[0][1]) + NanCheck(TransportUserInfo8[0][2])
            ws['F21'] = NanCheck(TransportUserInfo9[0][0]) + NanCheck(
                TransportUserInfo9[0][1]) + NanCheck(TransportUserInfo9[0][2])
            ws['F22'] = NanCheck(TransportUserInfo10[0][0]) + NanCheck(
                TransportUserInfo10[0][1]) + NanCheck(TransportUserInfo10[0][2])
            ws['F23'] = NanCheck(TransportUserInfo11[0][0]) + NanCheck(
                TransportUserInfo11[0][1]) + NanCheck(TransportUserInfo11[0][2])
            ws['F24'] = NanCheck(TransportUserInfo12[0][0]) + NanCheck(
                TransportUserInfo12[0][1]) + NanCheck(TransportUserInfo12[0][2])
            ws['F25'] = NanCheck(TransportUserInfo13[0][0]) + NanCheck(
                TransportUserInfo13[0][1]) + NanCheck(TransportUserInfo13[0][2])
            ws['F26'] = NanCheck(TransportUserInfo14[0][0]) + NanCheck(
                TransportUserInfo14[0][1]) + NanCheck(TransportUserInfo14[0][2])

            #Всего прибыло всеми видами транспорта
            ws['G20'] = NanCheck(TransportUserInfo8[0][4]) + NanCheck(
                TransportUserInfo8[0][5]) + NanCheck(TransportUserInfo8[0][6])
            ws['G21'] = NanCheck(TransportUserInfo9[0][4]) + NanCheck(
                TransportUserInfo9[0][5]) + NanCheck(TransportUserInfo9[0][6])
            ws['G22'] = NanCheck(TransportUserInfo10[0][4]) + NanCheck(
                TransportUserInfo10[0][5]) + NanCheck(TransportUserInfo10[0][6])
            ws['G23'] = NanCheck(TransportUserInfo11[0][4]) + NanCheck(
                TransportUserInfo11[0][5]) + NanCheck(TransportUserInfo11[0][6])
            ws['G24'] = NanCheck(TransportUserInfo12[0][4]) + NanCheck(
                TransportUserInfo12[0][5]) + NanCheck(TransportUserInfo12[0][6])
            ws['G25'] = NanCheck(TransportUserInfo13[0][4]) + NanCheck(
                TransportUserInfo13[0][5]) + NanCheck(TransportUserInfo13[0][6])
            ws['G26'] = NanCheck(TransportUserInfo14[0][4]) + NanCheck(
                TransportUserInfo14[0][5]) + NanCheck(TransportUserInfo14[0][6])


            ######3table

            #На рейде в т.ч на линейных судах В ожидании выгрузки
            ws['D32'] = NanCheck(Reid_info1[0][0])
            ws['D33'] = NanCheck(Reid_info2[0][0])
            ws['D34'] = NanCheck(Reid_info3[0][0])
            ws['D35'] = NanCheck(Reid_info4[0][0])
            ws['D36'] = NanCheck(Reid_info5[0][0])
            ws['D37'] = NanCheck(Reid_info6[0][0])
            ws['D38'] = NanCheck(Reid_info7[0][0])

            #На рейде в т.ч. на трамповых судах В ожидании выгрузки
            ws['E32'] = NanCheck(Reid_info1[0][1])
            ws['E33'] = NanCheck(Reid_info2[0][1])
            ws['E34'] = NanCheck(Reid_info3[0][1])
            ws['E35'] = NanCheck(Reid_info4[0][1])
            ws['E36'] = NanCheck(Reid_info5[0][1])
            ws['E37'] = NanCheck(Reid_info6[0][1])
            ws['E38'] = NanCheck(Reid_info7[0][1])

            #На рейде в т.ч на линейных судах В ожидании погрузки
            ws['G32'] = NanCheck(Reid_info1[0][2])
            ws['G33'] = NanCheck(Reid_info2[0][2])
            ws['G34'] = NanCheck(Reid_info3[0][2])
            ws['G35'] = NanCheck(Reid_info4[0][2])
            ws['G36'] = NanCheck(Reid_info5[0][2])
            ws['G37'] = NanCheck(Reid_info6[0][2])
            ws['G38'] = NanCheck(Reid_info7[0][2])


            #На рейде в т.ч. на трамповых судах В ожидании погрузки
            ws['H32'] = NanCheck(Reid_info1[0][3])
            ws['H33'] = NanCheck(Reid_info2[0][3])
            ws['H34'] = NanCheck(Reid_info3[0][3])
            ws['H35'] = NanCheck(Reid_info4[0][3])
            ws['H36'] = NanCheck(Reid_info5[0][3])
            ws['H37'] = NanCheck(Reid_info6[0][3])
            ws['H38'] = NanCheck(Reid_info7[0][3])

            #На рейде в т.ч на линейных судах В ожидании выгрузки
            ws['K32'] = NanCheck(Reid_info1[0][4])
            ws['K33'] = NanCheck(Reid_info2[0][4])
            ws['K34'] = NanCheck(Reid_info3[0][4])
            ws['K35'] = NanCheck(Reid_info4[0][4])
            ws['K36'] = NanCheck(Reid_info5[0][4])
            ws['K37'] = NanCheck(Reid_info6[0][4])
            ws['K38'] = NanCheck(Reid_info7[0][4])

            #На рейде в т.ч. на трамповых судах В ожидании выгрузки
            ws['L32'] = NanCheck(Reid_info1[0][5])
            ws['L33'] = NanCheck(Reid_info2[0][5])
            ws['L34'] = NanCheck(Reid_info3[0][5])
            ws['L35'] = NanCheck(Reid_info4[0][5])
            ws['L36'] = NanCheck(Reid_info5[0][5])
            ws['L37'] = NanCheck(Reid_info6[0][5])
            ws['L38'] = NanCheck(Reid_info7[0][5])

            #На рейде в т.ч на линейных судах В ожидании погрузки
            ws['N32'] = NanCheck(Reid_info1[0][6])
            ws['N33'] = NanCheck(Reid_info2[0][6])
            ws['N34'] = NanCheck(Reid_info3[0][6])
            ws['N35'] = NanCheck(Reid_info4[0][6])
            ws['N36'] = NanCheck(Reid_info5[0][6])
            ws['N37'] = NanCheck(Reid_info6[0][6])
            ws['N38'] = NanCheck(Reid_info7[0][6])


            #На рейде в т.ч. на трамповых судах В ожидании погрузки
            ws['O32'] = NanCheck(Reid_info1[0][7])
            ws['O33'] = NanCheck(Reid_info2[0][7])
            ws['O34'] = NanCheck(Reid_info3[0][7])
            ws['O35'] = NanCheck(Reid_info4[0][7])
            ws['O36'] = NanCheck(Reid_info5[0][7])
            ws['O37'] = NanCheck(Reid_info6[0][7])
            ws['O38'] = NanCheck(Reid_info7[0][7])


            ###4 table
            #Количество фитинговых платформ, ед. Всего на сети
            ws['C44'] = NanCheck(WagonsData1[0][0])
            ws['C45'] = NanCheck(WagonsData2[0][0])
            ws['C46'] = NanCheck(WagonsData3[0][0])
            ws['C47'] = NanCheck(WagonsData4[0][0])
            ws['C48'] = NanCheck(WagonsData5[0][0])
            ws['C49'] = NanCheck(WagonsData6[0][0])
            ws['C50'] = NanCheck(WagonsData7[0][0])

            #Количество фитинговых платформ, ед. в т.ч. на дальневосточной дороге

            ws['D44'] = NanCheck(WagonsDataFE1[0][0])
            ws['D45'] = NanCheck(WagonsDataFE2[0][0])
            ws['D46'] = NanCheck(WagonsDataFE3[0][0])
            ws['D47'] = NanCheck(WagonsDataFE4[0][0])
            ws['D48'] = NanCheck(WagonsDataFE5[0][0])
            ws['D49'] = NanCheck(WagonsDataFE6[0][0])
            ws['D50'] = NanCheck(WagonsDataFE7[0][0])


            ### 5 table
            #Убыло
            ws['C56'] = NanCheck(TransportUserInfo1[0][2])
            ws['C57'] = NanCheck(TransportUserInfo2[0][2])
            ws['C58'] = NanCheck(TransportUserInfo3[0][2])
            ws['C59'] = NanCheck(TransportUserInfo4[0][2])
            ws['C60'] = NanCheck(TransportUserInfo5[0][2])
            ws['C61'] = NanCheck(TransportUserInfo6[0][2])
            ws['C62'] = NanCheck(TransportUserInfo7[0][2])

            #Прибыло
            ws['D56'] = NanCheck(TransportUserInfo1[0][6])
            ws['D57'] = NanCheck(TransportUserInfo2[0][6])
            ws['D58'] = NanCheck(TransportUserInfo3[0][6])
            ws['D59'] = NanCheck(TransportUserInfo4[0][6])
            ws['D60'] = NanCheck(TransportUserInfo5[0][6])
            ws['D61'] = NanCheck(TransportUserInfo6[0][6])
            ws['D62'] = NanCheck(TransportUserInfo7[0][6])


            ####6 table
            #Наличие контейнеров в портовых терминалах, готовых к вывозу по ж.д.
            ws['C69'] = NanCheck(ContainerDataNow1[0][0])
            ws['C70'] = NanCheck(ContainerDataNow2[0][0])
            ws['C71'] = NanCheck(ContainerDataNow3[0][0])
            ws['C72'] = NanCheck(ContainerDataNow4[0][0])
            ws['C73'] = NanCheck(ContainerDataNow5[0][0])
            ws['C74'] = NanCheck(ContainerDataNow6[0][0])
            ws['C75'] = NanCheck(ContainerDataNow7[0][0])


            #Наличие контейнеров в портовых терминалах, готовых к вывозу автотранспортом
            ws['D69'] = NanCheck(ContainerDataNow1[0][1])
            ws['D70'] = NanCheck(ContainerDataNow2[0][1])
            ws['D71'] = NanCheck(ContainerDataNow3[0][1])
            ws['D72'] = NanCheck(ContainerDataNow4[0][1])
            ws['D73'] = NanCheck(ContainerDataNow5[0][1])
            ws['D74'] = NanCheck(ContainerDataNow6[0][1])
            ws['D75'] = NanCheck(ContainerDataNow7[0][1])

            # Обеспечение автотранспортом портовых терминалов
            ws['E69'] = NanCheck(ContainerDataNow1[0][2])
            ws['E70'] = NanCheck(ContainerDataNow2[0][2])
            ws['E71'] = NanCheck(ContainerDataNow3[0][2])
            ws['E72'] = NanCheck(ContainerDataNow4[0][2])
            ws['E73'] = NanCheck(ContainerDataNow5[0][2])
            ws['E74'] = NanCheck(ContainerDataNow6[0][2])
            ws['E75'] = NanCheck(ContainerDataNow7[0][2])
            conn.close()
        except:
            pass
        bytes_io = BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)
        return FileResponse(bytes_io,
                            as_attachment=True,
                            filename=(date+'.xlsx'))
    else:
        return redirect('home')

# show of dataset for datepick_admin-DAYDELTA date
@login_required(login_url='')
def dataset(request):
    if request.user.id == 1:
        conn = connection()
        params = request.session.get('parameters', {})
        date = params.get('date1')
        # try:
            #
            # TestUser1
            #
                #table 1
        Tranzit1 = getTranzitUserInfoFromDB(2,date,conn)
        User1data = getUserInfoFromDBDataset(2, date,conn)
        MaxWarehouseQty1User = getMaxWarehouseQty(2,conn)[0][0]
        NormsWarehouseQty1User = getNormsWarehouseQty(2,conn)[0][0]
        AllQty1User = AllQtyCalculator(User1data, Tranzit1)
        AllQtyPercent1User = AllQtyPercent(AllQty1User, MaxWarehouseQty1User, 0)
        Reid_info1 = getReidUserInfoFromDB(2,date,conn)
        ReidAllUser1 = ReidAllStr(Reid_info1)
        PortAllUser1 = PortAllStr(Reid_info1)
                #table 2
        ContainerData1 = getContaunerUserInfoFromDB(2,date,conn)
        ContainerDataNow1 = getContaunerUserInfoFromDB(2, (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime('%Y-%m-%d'),conn)
                #table3
        WagonsData1 = getWagonsUserInfoFromDB(2,date,conn)
                #table4
        WagonsDataFE1 = getWagonsUserInfoFromDBFE(2,date,conn)
                #table5
        TransportUserInfo1 = getTransportUserInfoFromDB(2,date,conn)
        TransportUserInfo1InPercent = TransportPercent((NanCheck(TransportUserInfo1[0][4])
                                                        + NanCheck(TransportUserInfo1[0][5])
                                                        + NanCheck(TransportUserInfo1[0][6])
                                                        + NanCheck(TransportUserInfo1[0][7])),
                                                       2400,
                                                       2)
        TransportUserInfo1OutPercent = TransportPercent((NanCheck(TransportUserInfo1[0][0])
                                                         + NanCheck(TransportUserInfo1[0][1])
                                                         + NanCheck(TransportUserInfo1[0][2])
                                                         + NanCheck(TransportUserInfo1[0][3])),
                                                        2400,
                                                        2)
        TransportCalculated1 = TransportCalculator(TransportUserInfo1)

        #
        # TestUser2
        #
                #table 1
        Tranzit2 = getTranzitUserInfoFromDB(3,date,conn)
        User2data = getUserInfoFromDBDataset(3, date,conn)
        MaxWarehouseQty2User = getMaxWarehouseQty(3,conn)[0][0]
        NormsWarehouseQty2User = getNormsWarehouseQty(3,conn)[0][0]
        AllQty2User = AllQtyCalculator(User2data, Tranzit2)
        AllQtyPercent2User = AllQtyPercent(AllQty2User, MaxWarehouseQty2User, 0)
        Reid_info2 = getReidUserInfoFromDB(3,date,conn)
        ReidAllUser2 = ReidAllStr(Reid_info2)
        PortAllUser2 = PortAllStr(Reid_info2)
                #table 2
        ContainerData2 = getContaunerUserInfoFromDB(3,date,conn)
        ContainerDataNow2 = getContaunerUserInfoFromDB(3,(dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime('%Y-%m-%d'),conn)
                #table3
        WagonsData2 = getWagonsUserInfoFromDB(3,date,conn)
                #table4
        WagonsDataFE2 = getWagonsUserInfoFromDBFE(3,date,conn)
                #table5
        TransportUserInfo2 = getTransportUserInfoFromDB(3,date,conn)
        TransportUserInfo2InPercent = TransportPercent((NanCheck(TransportUserInfo2[0][4])
                                                        + NanCheck(TransportUserInfo2[0][5])
                                                        + NanCheck(TransportUserInfo2[0][6])
                                                        + NanCheck(TransportUserInfo2[0][7])),
                                                       2860,
                                                       2)
        TransportUserInfo2OutPercent = TransportPercent((NanCheck(TransportUserInfo2[0][0])
                                                         + NanCheck(TransportUserInfo2[0][1])
                                                         + NanCheck(TransportUserInfo2[0][2])
                                                         + NanCheck(TransportUserInfo2[0][3])),
                                                        2860,
                                                        2)
        TransportCalculated2 = TransportCalculator(TransportUserInfo2)


        #
        # 3 user
        #

                #table 1
        Tranzit3 = getTranzitUserInfoFromDB(4,date,conn)
        User3data = getUserInfoFromDBDataset(4, date,conn)
        MaxWarehouseQty3User = getMaxWarehouseQty(4,conn)[0][0]
        NormsWarehouseQty3User = getNormsWarehouseQty(4,conn)[0][0]
        AllQty3User = AllQtyCalculator(User3data, Tranzit3)
        AllQtyPercent3User = AllQtyPercent(AllQty3User, MaxWarehouseQty3User, 0)
        Reid_info3 = getReidUserInfoFromDB(4,date,conn)
        ReidAllUser3 = ReidAllStr(Reid_info3)
        PortAllUser3 = PortAllStr(Reid_info3)
                #table 2
        ContainerData3 = getContaunerUserInfoFromDB(4,date,conn)
        ContainerDataNow3 = getContaunerUserInfoFromDB(4,(dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime('%Y-%m-%d'),conn)
                #table3
        WagonsData3 = getWagonsUserInfoFromDB(4,date,conn)
                #table4
        WagonsDataFE3 = getWagonsUserInfoFromDBFE(4,date,conn)
                #table5
        TransportUserInfo3 = getTransportUserInfoFromDB(4,date,conn)
        TransportUserInfo3InPercent = TransportPercent((NanCheck(TransportUserInfo3[0][4])
                                                        + NanCheck(TransportUserInfo3[0][5])
                                                        + NanCheck(TransportUserInfo3[0][6])
                                                        + NanCheck(TransportUserInfo3[0][7])),
                                                       800,
                                                       2)
        TransportUserInfo3OutPercent = TransportPercent((NanCheck(TransportUserInfo3[0][0])
                                                         + NanCheck(TransportUserInfo3[0][1])
                                                         + NanCheck(TransportUserInfo3[0][2])
                                                         + NanCheck(TransportUserInfo3[0][3])),
                                                        800,
                                                        2)
        TransportCalculated3 = TransportCalculator(TransportUserInfo3)

        #
        # 4 user
        #

        # table 1
        Tranzit4 = getTranzitUserInfoFromDB(5, date,conn)
        User4data = getUserInfoFromDBDataset(5, date,conn)
        MaxWarehouseQty4User = getMaxWarehouseQty(5,conn)[0][0]
        NormsWarehouseQty4User = getNormsWarehouseQty(5,conn)[0][0]
        AllQty4User = AllQtyCalculator(User4data, Tranzit4)
        AllQtyPercent4User = AllQtyPercent(AllQty4User, MaxWarehouseQty4User, 0)
        Reid_info4 = getReidUserInfoFromDB(5, date,conn)
        ReidAllUser4 = ReidAllStr(Reid_info4)
        PortAllUser4 = PortAllStr(Reid_info4)
        # table 2
        ContainerData4 = getContaunerUserInfoFromDB(5, date,conn)
        ContainerDataNow4 = getContaunerUserInfoFromDB(5, (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
            '%Y-%m-%d'),conn)
        # table3
        WagonsData4 = getWagonsUserInfoFromDB(5, date,conn)
        # table4
        WagonsDataFE4 = getWagonsUserInfoFromDBFE(5, date,conn)
        # table5
        TransportUserInfo4 = getTransportUserInfoFromDB(5, date,conn)
        TransportUserInfo4InPercent = TransportPercent((NanCheck(TransportUserInfo4[0][4])
                                                        + NanCheck(TransportUserInfo4[0][5])
                                                        + NanCheck(TransportUserInfo4[0][6])
                                                        + NanCheck(TransportUserInfo4[0][7])),
                                                       380,
                                                       2)
        TransportUserInfo4OutPercent = TransportPercent((NanCheck(TransportUserInfo4[0][0])
                                                         + NanCheck(TransportUserInfo4[0][1])
                                                         + NanCheck(TransportUserInfo4[0][2])
                                                         + NanCheck(TransportUserInfo4[0][3])),
                                                        380,
                                                        2)
        TransportCalculated4 = TransportCalculator(TransportUserInfo4)

        #
        # 5 user
        #

        # table 1
        Tranzit5 = getTranzitUserInfoFromDB(6, date,conn)
        User5data = getUserInfoFromDBDataset(6, date,conn)
        MaxWarehouseQty5User = getMaxWarehouseQty(6,conn)[0][0]
        NormsWarehouseQty5User = getNormsWarehouseQty(6,conn)[0][0]
        AllQty5User = AllQtyCalculator(User5data, Tranzit5)
        AllQtyPercent5User = AllQtyPercent(AllQty5User, MaxWarehouseQty5User, 0)
        Reid_info5 = getReidUserInfoFromDB(6, date,conn)
        ReidAllUser5 = ReidAllStr(Reid_info5)
        PortAllUser5 = PortAllStr(Reid_info5)
        # table 2
        ContainerData5 = getContaunerUserInfoFromDB(6, date,conn)
        ContainerDataNow5 = getContaunerUserInfoFromDB(6, (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
            '%Y-%m-%d'),conn)
        # table3
        WagonsData5 = getWagonsUserInfoFromDB(6, date,conn)
        # table4
        WagonsDataFE5 = getWagonsUserInfoFromDBFE(6, date,conn)
        # table5
        TransportUserInfo5 = getTransportUserInfoFromDB(6, date,conn)
        TransportUserInfo5InPercent = TransportPercent((NanCheck(TransportUserInfo5[0][4])
                                                        + NanCheck(TransportUserInfo5[0][5])
                                                        + NanCheck(TransportUserInfo5[0][6])
                                                        + NanCheck(TransportUserInfo5[0][7])),
                                                       580,
                                                       2)
        TransportUserInfo5OutPercent = TransportPercent((NanCheck(TransportUserInfo5[0][0])
                                                         + NanCheck(TransportUserInfo5[0][1])
                                                         + NanCheck(TransportUserInfo5[0][2])
                                                         + NanCheck(TransportUserInfo5[0][3])),
                                                        580,
                                                        2)
        TransportCalculated5 = TransportCalculator(TransportUserInfo5)

        #
        # 6 user
        #

        # table 1
        Tranzit6 = getTranzitUserInfoFromDB(7, date,conn)
        User6data = getUserInfoFromDBDataset(7, date,conn)
        MaxWarehouseQty6User = getMaxWarehouseQty(7,conn)[0][0]
        NormsWarehouseQty6User = getNormsWarehouseQty(7,conn)[0][0]
        AllQty6User = AllQtyCalculator(User6data, Tranzit6)
        AllQtyPercent6User = AllQtyPercent(AllQty6User, MaxWarehouseQty6User, 0)
        Reid_info6 = getReidUserInfoFromDB(7, date,conn)
        ReidAllUser6 = ReidAllStr(Reid_info6)
        PortAllUser6 = PortAllStr(Reid_info6)
        # table 2
        ContainerData6 = getContaunerUserInfoFromDB(7, date,conn)
        ContainerDataNow6 = getContaunerUserInfoFromDB(7, (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
            '%Y-%m-%d'),conn)
        # table3
        WagonsData6 = getWagonsUserInfoFromDB(7, date,conn)
        # table4
        WagonsDataFE6 = getWagonsUserInfoFromDBFE(7, date,conn)
        # table5
        TransportUserInfo6 = getTransportUserInfoFromDB(7, date,conn)
        TransportUserInfo6InPercent = TransportPercent((NanCheck(TransportUserInfo6[0][4])
                                                        + NanCheck(TransportUserInfo6[0][5])
                                                        + NanCheck(TransportUserInfo6[0][6])
                                                        + NanCheck(TransportUserInfo6[0][7])),
                                                       100,
                                                       2)
        TransportUserInfo6OutPercent = TransportPercent((NanCheck(TransportUserInfo6[0][0])
                                                         + NanCheck(TransportUserInfo6[0][1])
                                                         + NanCheck(TransportUserInfo6[0][2])
                                                         + NanCheck(TransportUserInfo6[0][3])),
                                                        100,
                                                        2)
        TransportCalculated6 = TransportCalculator(TransportUserInfo6)


        #
        # 7 user
        #

        # table 1
        Tranzit7 = getTranzitUserInfoFromDB(8, date,conn)
        User7data = getUserInfoFromDBDataset(8, date,conn)
        MaxWarehouseQty7User = getMaxWarehouseQty(8,conn)[0][0]
        NormsWarehouseQty7User = getNormsWarehouseQty(8,conn)[0][0]
        AllQty7User = AllQtyCalculator(User7data, Tranzit7)
        AllQtyPercent7User = AllQtyPercent(AllQty7User, MaxWarehouseQty7User, 0)
        Reid_info7 = getReidUserInfoFromDB(8, date,conn)
        ReidAllUser7 = ReidAllStr(Reid_info7)
        PortAllUser7 = PortAllStr(Reid_info7)
        # table 2
        ContainerData7 = getContaunerUserInfoFromDB(8, date,conn)
        ContainerDataNow7 = getContaunerUserInfoFromDB(8, (dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime(
            '%Y-%m-%d'),conn)
        # table3
        WagonsData7 = getWagonsUserInfoFromDB(8, date,conn)
        # table4
        WagonsDataFE7 = getWagonsUserInfoFromDBFE(8, date,conn)
        # table5
        TransportUserInfo7 = getTransportUserInfoFromDB(8, date,conn)
        TransportUserInfo7InPercent = TransportPercent((NanCheck(TransportUserInfo7[0][4])
                                                        + NanCheck(TransportUserInfo7[0][5])
                                                        + NanCheck(TransportUserInfo7[0][6])
                                                        + NanCheck(TransportUserInfo7[0][7])),
                                                       0,
                                                       2)
        TransportUserInfo7OutPercent = TransportPercent((NanCheck(TransportUserInfo7[0][0])
                                                         + NanCheck(TransportUserInfo7[0][1])
                                                         + NanCheck(TransportUserInfo7[0][2])
                                                         + NanCheck(TransportUserInfo7[0][3])),
                                                        0,
                                                        2)
        TransportCalculated7 = TransportCalculator(TransportUserInfo7)


        ###########
        # AllUsers#
        ###########

            #table 1
        UserAlldata = getDataTableForAllTime(date,conn)
        TranzitAll = getAllTranzitUserInfoFromDB(date,conn)
        MaxWarehouseQtyAll = getMaxWarehouseAllQty(conn)[0][0]
        NormsWarehouseQtyAll = getNormsWarehouseAllQty(conn)[0][0]
        AllQtyAll = AllQtyCalculator(UserAlldata, TranzitAll)
        AllQtyPercent1All = AllQtyPercent(AllQtyAll, MaxWarehouseQtyAll, 0)
        Reid_infoAll = getReidAllInfoFromDB(date,conn)
        ReidAllAll = ReidAllStr(Reid_infoAll)
        PortAllAll = PortAllStr(Reid_infoAll)
            #table 2
        ContainerDataAll = getContaunerInfoFromDBAll(date,conn)
        ContainerDataAllNow = getContaunerInfoFromDBAll((dt.datetime.strptime(date, '%Y-%m-%d') + DAYDELTA).strftime('%Y-%m-%d'),conn)
            # table3
        WagonDataAll = getWagonsInfoFromDBAll(date,conn)
            # table4
        WagonDataAllFE = getWagonsInfoFromDBAllFE(date,conn)
                #table5
        TransportUserInfoAll = getTransportInfoFromDBAll(date,conn)
        TransportUserInfoAllInPercent = TransportPercent((NanCheck(TransportUserInfoAll[0][4])
                                                          + NanCheck(TransportUserInfoAll[0][5])
                                                          + NanCheck(TransportUserInfoAll[0][6])
                                                          + NanCheck(TransportUserInfoAll[0][7])),
                                                         7120,
                                                         2)
        TransportUserInfoAllOutPercent = TransportPercent((NanCheck(TransportUserInfoAll[0][0])
                                                           + NanCheck(TransportUserInfoAll[0][1])
                                                           + NanCheck(TransportUserInfoAll[0][2])
                                                           + NanCheck(TransportUserInfoAll[0][3])),
                                                          7120,
                                                          2)
        TransportCalculatedAll = TransportCalculator(TransportUserInfoAll)



        #####
        #######6 table##### 2 table in dataset
        #####

        # 8 user

        TransportUserInfo8 = getTransportUserInfoFromDB(9, date,conn)
        TransportUserInfo8Percent = TransportPercent(NanCheck(TransportUserInfo8[0][8]),
                                                     NanCheck(getMaxWarehouseQty(9,conn)[0][0]),
                                                     2)
        TransportCalculated8 = TransportCalculator(TransportUserInfo8)

        # 9 user

        TransportUserInfo9 = getTransportUserInfoFromDB(10, date,conn)
        TransportUserInfo9Percent = TransportPercent(NanCheck(TransportUserInfo9[0][8]),
                                                     NanCheck(getMaxWarehouseQty(10,conn)[0][0]),
                                                     2)
        TransportCalculated9 = TransportCalculator(TransportUserInfo9)

        # 10 user

        TransportUserInfo10 = getTransportUserInfoFromDB(11, date,conn)
        TransportUserInfo10Percent = TransportPercent(NanCheck(TransportUserInfo10[0][8]),
                                                      NanCheck(getMaxWarehouseQty(11,conn)[0][0]),
                                                      2)
        TransportCalculated10 = TransportCalculator(TransportUserInfo10)

        # 11 user

        TransportUserInfo11 = getTransportUserInfoFromDB(12, date,conn)
        TransportUserInfo11Percent = TransportPercent(NanCheck(TransportUserInfo11[0][8]),
                                                      NanCheck(getMaxWarehouseQty(12,conn)[0][0]),
                                                      2)
        TransportCalculated11 = TransportCalculator(TransportUserInfo11)

        # 12 user

        TransportUserInfo12 = getTransportUserInfoFromDB(13, date,conn)
        TransportUserInfo12Percent = TransportPercent(NanCheck(TransportUserInfo12[0][8]),
                                                      NanCheck(getMaxWarehouseQty(13,conn)[0][0]),
                                                      2)
        TransportCalculated12 = TransportCalculator(TransportUserInfo12)

        # 13 user

        TransportUserInfo13 = getTransportUserInfoFromDB(14, date,conn)
        TransportUserInfo13Percent = TransportPercent(NanCheck(TransportUserInfo13[0][8]),
                                                      NanCheck(getMaxWarehouseQty(14,conn)[0][0]),
                                                      2)
        TransportCalculated13 = TransportCalculator(TransportUserInfo13)

        # 14 user

        TransportUserInfo14 = getTransportUserInfoFromDB(15, date,conn)
        TransportUserInfo14Percent = TransportPercent(NanCheck(TransportUserInfo14[0][8]),
                                                      NanCheck(getMaxWarehouseQty(15,conn)[0][0]),
                                                      2)
        TransportCalculated14 = TransportCalculator(TransportUserInfo14)

        MaxWarehouseQty8User = NanCheck(getMaxWarehouseQty(9, conn)[0][0])
        MaxWarehouseQty9User = NanCheck(getMaxWarehouseQty(10, conn)[0][0])
        MaxWarehouseQty10User = NanCheck(getMaxWarehouseQty(11, conn)[0][0])
        MaxWarehouseQty11User = NanCheck(getMaxWarehouseQty(12, conn)[0][0])
        MaxWarehouseQty12User = NanCheck(getMaxWarehouseQty(13, conn)[0][0])
        MaxWarehouseQty13User = NanCheck(getMaxWarehouseQty(14, conn)[0][0])
        MaxWarehouseQty14User = NanCheck(getMaxWarehouseQty(15, conn)[0][0])
        MaxWarehouseQtyAllNotST = NanCheck(getMaxWarehouseAllQtyNotST(conn)[0][0])
        #
        #All not stevedor users
        #

        TransportUserInfoNotSTAll = getTransportInfoFromDBNotSTAll(date,conn)

        TransportUserInfoAllPercent = TransportPercent(NanCheck(TransportUserInfoNotSTAll[0][8]),
                                                       NanCheck(getMaxWarehouseAllQtyNotST(conn)[0][0]),
                                                       2)
        TransportCalculatedAllNotST = TransportCalculator(TransportUserInfoNotSTAll)
        conn.close()
        # except:
        #     return render(request, 'error.html', {'ErrorText' : 'Ошибка отображения данных'})
        return render(request, 'dataset.html', {
                                                    'date' : ((dt.datetime.strptime(date,'%Y-%m-%d')+DAYDELTA).strftime('%d.%m.%Y.')),
            # Dataset for all tables
                                                    'User1data' : User1data[0],
                                                    'User2data': User2data[0],
                                                    'UserAlldata': UserAlldata[0],

                            # 1 User
                                                    'AllQtyPercent1User': AllQtyPercent1User,
                                                    'AllQty1User' :AllQty1User,
                                                    'MaxWarehouseQty1User': MaxWarehouseQty1User,
                                                    'NormsWarehouseQty1User' : NormsWarehouseQty1User,
                                                    'ContainerData1': NanCheck(ContainerData1[0][0]),
                                                    'ContainerDataNow1_0': NanCheck(ContainerDataNow1[0][0]),
                                                    'ContainerDataNow1_1': NanCheck(ContainerDataNow1[0][1]),
                                                    'ContainerDataNow1_2': NanCheck(ContainerDataNow1[0][2]),
                                                    'WagonsData1': NanCheck(WagonsData1[0][0]),
                                                    'WagonsDataQty1': NanCheck(WagonsData1[0][0]) * 3.5,
                                                    'WagonsDataFE1': NanCheck(WagonsDataFE1[0][0]),
                                                    'WagonsDataQtyFE1': NanCheck(WagonsDataFE1[0][0]) * 3.5,
                                                    'TransportUserInfo1In' : NanCheck(TransportUserInfo1[0][4])
                                                                             + NanCheck(TransportUserInfo1[0][5])
                                                                             + NanCheck(TransportUserInfo1[0][6])
                                                                             + NanCheck(TransportUserInfo1[0][7]),
                                                    'TransportUserInfo1InPercent' : TransportUserInfo1InPercent,
                                                    'TransportUserInfo1Out': NanCheck(TransportUserInfo1[0][0])
                                                                             + NanCheck(TransportUserInfo1[0][1])
                                                                             + NanCheck(TransportUserInfo1[0][2])
                                                                             + NanCheck(TransportUserInfo1[0][3]),
                                                    'TransportUserInfo1OutPercent': TransportUserInfo1OutPercent,
                                                    'TransportCalculated1':TransportCalculated1,
                                                    'AutoTransportIn1': NanCheck(TransportUserInfo1[0][6]),
                                                    'AutoTransportOut1': NanCheck(TransportUserInfo1[0][2]),
                                                    'AutoTransportCalculated1': AutoCalculator(TransportUserInfo1),



                            # 2 user
                                                    'AllQtyPercent2User': AllQtyPercent2User,
                                                    'AllQty2User': AllQty2User,
                                                    'MaxWarehouseQty2User': MaxWarehouseQty2User,
                                                    'NormsWarehouseQty2User': NormsWarehouseQty2User,
                                                    'ContainerData2': NanCheck(ContainerData2[0][0]),
                                                    'ContainerDataNow2_0': NanCheck(ContainerDataNow2[0][0]),
                                                    'ContainerDataNow2_1': NanCheck(ContainerDataNow2[0][1]),
                                                    'ContainerDataNow2_2': NanCheck(ContainerDataNow2[0][2]),
                                                    'WagonsData2': NanCheck(WagonsData2[0][0]),
                                                    'WagonsDataQty2': NanCheck(WagonsData2[0][0]) * 3.5,
                                                    'WagonsDataFE2': NanCheck(WagonsDataFE2[0][0]),
                                                    'WagonsDataQtyFE2': NanCheck(WagonsDataFE2[0][0]) * 3.5,
                                                    'TransportUserInfo2In': NanCheck(TransportUserInfo2[0][4])
                                                                            + NanCheck(TransportUserInfo2[0][5])
                                                                            + NanCheck(TransportUserInfo2[0][6])
                                                                            + NanCheck(TransportUserInfo2[0][7]),
                                                    'TransportUserInfo2InPercent': TransportUserInfo2InPercent,
                                                    'TransportUserInfo2Out': NanCheck(TransportUserInfo2[0][0])
                                                                             + NanCheck(TransportUserInfo2[0][1])
                                                                             + NanCheck(TransportUserInfo2[0][2])
                                                                             + NanCheck(TransportUserInfo2[0][3]),
                                                    'TransportUserInfo2OutPercent': TransportUserInfo2OutPercent,
                                                    'TransportCalculated2':TransportCalculated2,
                                                    'AutoTransportIn2': NanCheck(TransportUserInfo2[0][6]),
                                                    'AutoTransportOut2': NanCheck(TransportUserInfo2[0][2]),
                                                    'AutoTransportCalculated2': AutoCalculator(TransportUserInfo2),

                            #3 user


                                                    'AllQtyPercent3User': AllQtyPercent3User,
                                                    'AllQty3User': AllQty3User,
                                                    'MaxWarehouseQty3User': MaxWarehouseQty3User,
                                                    'NormsWarehouseQty3User': NormsWarehouseQty3User,
                                                    'ContainerData3': NanCheck(ContainerData3[0][0]),
                                                    'ContainerDataNow3_0': NanCheck(ContainerDataNow3[0][0]),
                                                    'ContainerDataNow3_1': NanCheck(ContainerDataNow3[0][1]),
                                                    'ContainerDataNow3_2': NanCheck(ContainerDataNow3[0][2]),
                                                    'WagonsData3': NanCheck(WagonsData3[0][0]),
                                                    'WagonsDataQty3': NanCheck(WagonsData3[0][0]) * 3.5,
                                                    'WagonsDataFE3': NanCheck(WagonsDataFE3[0][0]),
                                                    'WagonsDataQtyFE3': NanCheck(WagonsDataFE3[0][0]) * 3.5,
                                                    'TransportUserInfo3In': NanCheck(TransportUserInfo3[0][4])
                                                                            + NanCheck(TransportUserInfo3[0][5])
                                                                            + NanCheck(TransportUserInfo3[0][6])
                                                                            + NanCheck(TransportUserInfo3[0][7]),
                                                    'TransportUserInfo3InPercent': TransportUserInfo3InPercent,
                                                    'TransportUserInfo3Out': NanCheck(TransportUserInfo3[0][0])
                                                                             + NanCheck(TransportUserInfo3[0][1])
                                                                             + NanCheck(TransportUserInfo3[0][2])
                                                                             + NanCheck(TransportUserInfo3[0][3]),
                                                    'TransportUserInfo3OutPercent': TransportUserInfo3OutPercent,
                                                    'TransportCalculated3': TransportCalculated3,
                                                    'AutoTransportIn3': NanCheck(TransportUserInfo3[0][6]),
                                                    'AutoTransportOut3': NanCheck(TransportUserInfo3[0][2]),
                                                    'AutoTransportCalculated3': AutoCalculator(TransportUserInfo3),

                            # 4 user

                                                    'AllQtyPercent4User': AllQtyPercent4User,
                                                    'AllQty4User': AllQty4User,
                                                    'MaxWarehouseQty4User': MaxWarehouseQty4User,
                                                    'NormsWarehouseQty4User': NormsWarehouseQty4User,
                                                    'ContainerData4': NanCheck(ContainerData4[0][0]),
                                                    'ContainerDataNow4_0': NanCheck(ContainerDataNow4[0][0]),
                                                    'ContainerDataNow4_1': NanCheck(ContainerDataNow4[0][1]),
                                                    'ContainerDataNow4_2': NanCheck(ContainerDataNow4[0][2]),
                                                    'WagonsData4': NanCheck(WagonsData4[0][0]),
                                                    'WagonsDataQty4': NanCheck(WagonsData4[0][0]) * 3.5,
                                                    'WagonsDataFE4': NanCheck(WagonsDataFE4[0][0]),
                                                    'WagonsDataQtyFE4': NanCheck(WagonsDataFE4[0][0]) * 3.5,
                                                    'TransportUserInfo4In': NanCheck(TransportUserInfo4[0][4])
                                                                            + NanCheck(TransportUserInfo4[0][5])
                                                                            + NanCheck(TransportUserInfo4[0][6])
                                                                            + NanCheck(TransportUserInfo4[0][7]),
                                                    'TransportUserInfo4InPercent': TransportUserInfo4InPercent,
                                                    'TransportUserInfo4Out': NanCheck(TransportUserInfo4[0][0])
                                                                             + NanCheck(TransportUserInfo4[0][1])
                                                                             + NanCheck(TransportUserInfo4[0][2])
                                                                             + NanCheck(TransportUserInfo4[0][3]),
                                                    'TransportUserInfo4OutPercent': TransportUserInfo4OutPercent,
                                                    'TransportCalculated4': TransportCalculated4,
                                                    'AutoTransportIn4': NanCheck(TransportUserInfo4[0][6]),
                                                    'AutoTransportOut4': NanCheck(TransportUserInfo4[0][2]),
                                                    'AutoTransportCalculated4': AutoCalculator(TransportUserInfo4),

                            # 5 user

                                                    'AllQtyPercent5User': AllQtyPercent5User,
                                                    'AllQty5User': AllQty5User,
                                                    'MaxWarehouseQty5User': MaxWarehouseQty5User,
                                                    'NormsWarehouseQty5User': NormsWarehouseQty5User,
                                                    'ContainerData5': NanCheck(ContainerData5[0][0]),
                                                    'ContainerDataNow5_0': NanCheck(ContainerDataNow5[0][0]),
                                                    'ContainerDataNow5_1': NanCheck(ContainerDataNow5[0][1]),
                                                    'ContainerDataNow5_2': NanCheck(ContainerDataNow5[0][2]),
                                                    'WagonsData5': NanCheck(WagonsData5[0][0]),
                                                    'WagonsDataQty5': NanCheck(WagonsData5[0][0]) * 3.5,
                                                    'WagonsDataFE5': NanCheck(WagonsDataFE5[0][0]),
                                                    'WagonsDataQtyFE5': NanCheck(WagonsDataFE5[0][0]) * 3.5,
                                                    'TransportUserInfo5In': NanCheck(TransportUserInfo5[0][4])
                                                                            + NanCheck(TransportUserInfo5[0][5])
                                                                            + NanCheck(TransportUserInfo5[0][6])
                                                                            + NanCheck(TransportUserInfo5[0][7]),
                                                    'TransportUserInfo5InPercent': TransportUserInfo5InPercent,
                                                    'TransportUserInfo5Out': NanCheck(TransportUserInfo5[0][0])
                                                                             + NanCheck(TransportUserInfo5[0][1])
                                                                             + NanCheck(TransportUserInfo5[0][2])
                                                                             + NanCheck(TransportUserInfo5[0][3]),
                                                    'TransportUserInfo5OutPercent': TransportUserInfo5OutPercent,
                                                    'TransportCalculated5': TransportCalculated5,
                                                    'AutoTransportIn5': NanCheck(TransportUserInfo5[0][6]),
                                                    'AutoTransportOut5': NanCheck(TransportUserInfo5[0][2]),
                                                    'AutoTransportCalculated5': AutoCalculator(TransportUserInfo5),

                            # 6 user

                                                    'AllQtyPercent6User': AllQtyPercent6User,
                                                    'AllQty6User': AllQty6User,
                                                    'MaxWarehouseQty6User': MaxWarehouseQty6User,
                                                    'NormsWarehouseQty6User': NormsWarehouseQty6User,
                                                    'ContainerData6': NanCheck(ContainerData6[0][0]),
                                                    'ContainerDataNow6_0': NanCheck(ContainerDataNow6[0][0]),
                                                    'ContainerDataNow6_1': NanCheck(ContainerDataNow6[0][1]),
                                                    'ContainerDataNow6_2': NanCheck(ContainerDataNow6[0][2]),
                                                    'WagonsData6': NanCheck(WagonsData6[0][0]),
                                                    'WagonsDataQty6': NanCheck(WagonsData6[0][0]) * 3.5,
                                                    'WagonsDataFE6': NanCheck(WagonsDataFE6[0][0]),
                                                    'WagonsDataQtyFE6': NanCheck(WagonsDataFE6[0][0]) * 3.5,
                                                    'TransportUserInfo6In': NanCheck(TransportUserInfo6[0][4])
                                                                            + NanCheck(TransportUserInfo6[0][5])
                                                                            + NanCheck(TransportUserInfo6[0][6])
                                                                            + NanCheck(TransportUserInfo6[0][7]),
                                                    'TransportUserInfo6InPercent': TransportUserInfo6InPercent,
                                                    'TransportUserInfo6Out': NanCheck(TransportUserInfo6[0][0])
                                                                             + NanCheck(TransportUserInfo6[0][1])
                                                                             + NanCheck(TransportUserInfo6[0][2])
                                                                             + NanCheck(TransportUserInfo6[0][3]),
                                                    'TransportUserInfo6OutPercent': TransportUserInfo6OutPercent,
                                                    'TransportCalculated6': TransportCalculated6,
                                                    'AutoTransportIn6': NanCheck(TransportUserInfo6[0][6]),
                                                    'AutoTransportOut6': NanCheck(TransportUserInfo6[0][2]),
                                                    'AutoTransportCalculated6': AutoCalculator(TransportUserInfo6),

                                # 7 user

                                                    'AllQtyPercent7User': AllQtyPercent7User,
                                                    'AllQty7User': AllQty7User,
                                                    'MaxWarehouseQty7User': MaxWarehouseQty7User,
                                                    'NormsWarehouseQty7User': NormsWarehouseQty7User,
                                                    'ContainerData7': NanCheck(ContainerData7[0][0]),
                                                    'ContainerDataNow7_0': NanCheck(ContainerDataNow7[0][0]),
                                                    'ContainerDataNow7_1': NanCheck(ContainerDataNow7[0][1]),
                                                    'ContainerDataNow7_2': NanCheck(ContainerDataNow7[0][2]),
                                                    'WagonsData7': NanCheck(WagonsData7[0][0]),
                                                    'WagonsDataQty7': NanCheck(WagonsData7[0][0]) * 3.5,
                                                    'WagonsDataFE7': NanCheck(WagonsDataFE7[0][0]),
                                                    'WagonsDataQtyFE7': NanCheck(WagonsDataFE7[0][0]) * 3.5,
                                                    'TransportUserInfo7In': NanCheck(TransportUserInfo7[0][4])
                                                                            + NanCheck(TransportUserInfo7[0][5])
                                                                            + NanCheck(TransportUserInfo7[0][6])
                                                                            + NanCheck(TransportUserInfo7[0][7]),
                                                    'TransportUserInfo7InPercent': TransportUserInfo7InPercent,
                                                    'TransportUserInfo7Out': NanCheck(TransportUserInfo7[0][0])
                                                                             + NanCheck(TransportUserInfo7[0][1])
                                                                             + NanCheck(TransportUserInfo7[0][2])
                                                                             + NanCheck(TransportUserInfo7[0][3]),
                                                    'TransportUserInfo7OutPercent': TransportUserInfo7OutPercent,
                                                    'TransportCalculated7': TransportCalculated7,
                                                    'AutoTransportIn7': NanCheck(TransportUserInfo7[0][6]),
                                                    'AutoTransportOut7': NanCheck(TransportUserInfo7[0][2]),
                                                    'AutoTransportCalculated7': AutoCalculator(TransportUserInfo7),

            #alluser
                                                    'AllQtyAll' :AllQtyAll,
                                                    'AllQtyPercent1All' : AllQtyPercent1All,
                                                    'MaxWarehouseQtyAll':MaxWarehouseQtyAll,
                                                    'NormsWarehouseQtyAll':NormsWarehouseQtyAll,
                                                    'ContainerDataAll': NanCheck(ContainerDataAll[0][0]),
                                                    'ContainerDataAllNow_0': NanCheck(ContainerDataAllNow[0][0]),
                                                    'ContainerDataAllNow_1': NanCheck(ContainerDataAllNow[0][1]),
                                                    'ContainerDataAllNow_2': NanCheck(ContainerDataAllNow[0][2]),
                                                    'WagonsDataAll': NanCheck(WagonDataAll[0][0]),
                                                    'WagonsDataQtyAll': NanCheck(WagonDataAll[0][0]) * 3.5,
                                                    'WagonsDataAllFE': NanCheck(WagonDataAllFE[0][0]),
                                                    'WagonsDataQtyAllFE': NanCheck(WagonDataAllFE[0][0]) * 3.5,
                                                    'TransportUserInfoAll': NanCheck(TransportUserInfoAll[0][4])
                                                                            + NanCheck(TransportUserInfoAll[0][5])
                                                                            + NanCheck(TransportUserInfoAll[0][6])
                                                                            + NanCheck(TransportUserInfoAll[0][7]),
                                                    'TransportUserInfoAllInPercent': TransportUserInfoAllInPercent,
                                                    'TransportUserInfoAllOut': NanCheck(TransportUserInfoAll[0][0])
                                                                               + NanCheck(TransportUserInfoAll[0][1])
                                                                               + NanCheck(TransportUserInfoAll[0][2])
                                                                               + NanCheck(TransportUserInfoAll[0][3]),
                                                    'TransportUserInfoAllOutPercent': TransportUserInfoAllOutPercent,
                                                    'TransportCalculatedAll':TransportCalculatedAll,
                                                    'AutoTransportInAll': NanCheck(TransportUserInfoAll[0][6]),
                                                    'AutoTransportOutAll': NanCheck(TransportUserInfoAll[0][2]),
                                                    'AutoTransportCalculatedAll': AutoCalculator(TransportUserInfoAll),

            # 3 table dataset


                        # 1 user
                                                    'UnloadReidlin1User': NanCheck(Reid_info1[0][0]),
                                                    'UnloadReidtramp1User': NanCheck(Reid_info1[0][1]),
                                                    'UnloadReidSum1User' : (
                                                                NanCheck(Reid_info1[0][0]) + NanCheck(Reid_info1[0][1])),
                                                    'LoadingReidLin1User': NanCheck(Reid_info1[0][2]),
                                                    'LoadingReidTramp1User': NanCheck(Reid_info1[0][3]),
                                                    'LoadingReidSum1User' :(
                                                                NanCheck(Reid_info1[0][2]) + NanCheck(Reid_info1[0][3])),
                                                    'ReidAllUser1' : NanCheck(ReidAllUser1),
                                                    'UnloadPortlin1User' : NanCheck(Reid_info1[0][4]),
                                                    'UnloadPorttramp1User': NanCheck(Reid_info1[0][5]),
                                                    'UnloadPortSum1User' : (
                                                                NanCheck(Reid_info1[0][4]) + NanCheck(Reid_info1[0][5])),
                                                    'LoadingPortLin1User': NanCheck(Reid_info1[0][6]),
                                                    'LoadingPortTramp1User': NanCheck(Reid_info1[0][7]),
                                                    'LoadingPortSum1User': (
                                                                NanCheck(Reid_info1[0][6]) + NanCheck(Reid_info1[0][7])),
                                                    'PortAllUser1':PortAllUser1,

                    # 2 user
                                                    'UnloadReidlin2User': NanCheck(Reid_info2[0][0]),
                                                    'UnloadReidtramp2User': NanCheck(Reid_info2[0][1]),
                                                    'UnloadReidSum2User': (
                                                                NanCheck(Reid_info2[0][0]) + NanCheck(Reid_info2[0][1])),
                                                    'LoadingReidLin2User': NanCheck(Reid_info2[0][2]),
                                                    'LoadingReidTramp2User': NanCheck(Reid_info2[0][3]),
                                                    'LoadingReidSum2User': (
                                                                NanCheck(Reid_info2[0][2]) + NanCheck(Reid_info2[0][3])),
                                                    'ReidAllUser2': ReidAllUser2,
                                                    'UnloadPortlin2User': NanCheck(Reid_info2[0][4]),
                                                    'UnloadPorttramp2User': NanCheck(Reid_info2[0][5]),
                                                    'UnloadPortSum2User': (
                                                                NanCheck(Reid_info2[0][4]) + NanCheck(Reid_info2[0][5])),
                                                    'LoadingPortLin2User': NanCheck(Reid_info2[0][6]),
                                                    'LoadingPortTramp2User': NanCheck(Reid_info2[0][7]),
                                                    'LoadingPortSum2User': (
                                                                NanCheck(Reid_info2[0][6]) + NanCheck(Reid_info2[0][7])),
                                                    'PortAllUser2': PortAllUser2,

                    # 3 user
                                                    'UnloadReidlin3User': NanCheck(Reid_info3[0][0]),
                                                    'UnloadReidtramp3User': NanCheck(Reid_info3[0][1]),
                                                    'UnloadReidSum3User': (
                                                                NanCheck(Reid_info3[0][0]) + NanCheck(Reid_info3[0][1])),
                                                    'LoadingReidLin3User': NanCheck(Reid_info3[0][2]),
                                                    'LoadingReidTramp3User': NanCheck(Reid_info3[0][3]),
                                                    'LoadingReidSum3User': (
                                                                NanCheck(Reid_info3[0][2]) + NanCheck(Reid_info3[0][3])),
                                                    'ReidAllUser3': ReidAllUser3,
                                                    'UnloadPortlin3User': NanCheck(Reid_info3[0][4]),
                                                    'UnloadPorttramp3User': NanCheck(Reid_info3[0][5]),
                                                    'UnloadPortSum3User': (
                                                                NanCheck(Reid_info3[0][4]) + NanCheck(Reid_info3[0][5])),
                                                    'LoadingPortLin3User': NanCheck(Reid_info3[0][6]),
                                                    'LoadingPortTramp3User': NanCheck(Reid_info3[0][7]),
                                                    'LoadingPortSum3User': (
                                                                NanCheck(Reid_info3[0][6]) + NanCheck(Reid_info3[0][7])),
                                                    'PortAllUser3': PortAllUser3,

                    # 4 user
                                                    'UnloadReidlin4User': NanCheck(Reid_info4[0][0]),
                                                    'UnloadReidtramp4User': NanCheck(Reid_info4[0][1]),
                                                    'UnloadReidSum4User': (
                                                                NanCheck(Reid_info4[0][0]) + NanCheck(Reid_info4[0][1])),
                                                    'LoadingReidLin4User': NanCheck(Reid_info4[0][2]),
                                                    'LoadingReidTramp4User': NanCheck(Reid_info4[0][3]),
                                                    'LoadingReidSum4User': (
                                                                NanCheck(Reid_info4[0][2]) + NanCheck(Reid_info4[0][3])),
                                                    'ReidAllUser4': ReidAllUser4,
                                                    'UnloadPortlin4User': NanCheck(Reid_info4[0][4]),
                                                    'UnloadPorttramp4User': NanCheck(Reid_info4[0][5]),
                                                    'UnloadPortSum4User': (
                                                                NanCheck(Reid_info4[0][4]) + NanCheck(Reid_info4[0][5])),
                                                    'LoadingPortLin4User': NanCheck(Reid_info4[0][6]),
                                                    'LoadingPortTramp4User': NanCheck(Reid_info4[0][7]),
                                                    'LoadingPortSum4User': (
                                                                NanCheck(Reid_info4[0][6]) + NanCheck(Reid_info4[0][7])),
                                                    'PortAllUser4': PortAllUser4,

                    # 5 user
                                                    'UnloadReidlin5User': NanCheck(Reid_info5[0][0]),
                                                    'UnloadReidtramp5User': NanCheck(Reid_info5[0][1]),
                                                    'UnloadReidSum5User': (
                                                                NanCheck(Reid_info5[0][0]) + NanCheck(Reid_info5[0][1])),
                                                    'LoadingReidLin5User': NanCheck(Reid_info5[0][2]),
                                                    'LoadingReidTramp5User': NanCheck(Reid_info5[0][3]),
                                                    'LoadingReidSum5User': (
                                                                NanCheck(Reid_info5[0][2]) + NanCheck(Reid_info5[0][3])),
                                                    'ReidAllUser5': ReidAllUser5,
                                                    'UnloadPortlin5User': NanCheck(Reid_info5[0][4]),
                                                    'UnloadPorttramp5User': NanCheck(Reid_info5[0][5]),
                                                    'UnloadPortSum5User': (
                                                                NanCheck(Reid_info5[0][4]) + NanCheck(Reid_info5[0][5])),
                                                    'LoadingPortLin5User': NanCheck(Reid_info5[0][6]),
                                                    'LoadingPortTramp5User': NanCheck(Reid_info5[0][7]),
                                                    'LoadingPortSum5User': (
                                                                NanCheck(Reid_info5[0][6]) + NanCheck(Reid_info5[0][7])),
                                                    'PortAllUser5': PortAllUser5,

                    # 6 user
                                                    'UnloadReidlin6User': NanCheck(Reid_info6[0][0]),
                                                    'UnloadReidtramp6User': NanCheck(Reid_info6[0][1]),
                                                    'UnloadReidSum6User': (
                                                                NanCheck(Reid_info6[0][0]) + NanCheck(Reid_info6[0][1])),
                                                    'LoadingReidLin6User': NanCheck(Reid_info6[0][2]),
                                                    'LoadingReidTramp6User': NanCheck(Reid_info6[0][3]),
                                                    'LoadingReidSum6User': (
                                                                NanCheck(Reid_info6[0][2]) + NanCheck(Reid_info6[0][3])),
                                                    'ReidAllUser6': ReidAllUser6,
                                                    'UnloadPortlin6User': NanCheck(Reid_info6[0][4]),
                                                    'UnloadPorttramp6User': NanCheck(Reid_info6[0][5]),
                                                    'UnloadPortSum6User': (
                                                                NanCheck(Reid_info6[0][4]) + NanCheck(Reid_info6[0][5])),
                                                    'LoadingPortLin6User': NanCheck(Reid_info6[0][6]),
                                                    'LoadingPortTramp6User': NanCheck(Reid_info6[0][7]),
                                                    'LoadingPortSum6User': (
                                                                NanCheck(Reid_info6[0][6]) + NanCheck(Reid_info6[0][7])),
                                                    'PortAllUser6': PortAllUser6,

                    # 6 user
                                                    'UnloadReidlin7User': NanCheck(Reid_info7[0][0]),
                                                    'UnloadReidtramp7User': NanCheck(Reid_info7[0][1]),
                                                    'UnloadReidSum7User': (
                                                                NanCheck(Reid_info7[0][0]) + NanCheck(Reid_info7[0][1])),
                                                    'LoadingReidLin7User': NanCheck(Reid_info7[0][2]),
                                                    'LoadingReidTramp7User': NanCheck(Reid_info7[0][3]),
                                                    'LoadingReidSum7User': (
                                                                NanCheck(Reid_info7[0][2]) + NanCheck(Reid_info7[0][3])),
                                                    'ReidAllUser7': ReidAllUser7,
                                                    'UnloadPortlin7User': NanCheck(Reid_info7[0][4]),
                                                    'UnloadPorttramp7User': NanCheck(Reid_info7[0][5]),
                                                    'UnloadPortSum7User': (
                                                                NanCheck(Reid_info7[0][4]) + NanCheck(Reid_info7[0][5])),
                                                    'LoadingPortLin7User': NanCheck(Reid_info7[0][6]),
                                                    'LoadingPortTramp7User': NanCheck(Reid_info7[0][7]),
                                                    'LoadingPortSum7User': (
                                                                NanCheck(Reid_info7[0][6]) + NanCheck(Reid_info7[0][7])),
                                                    'PortAllUser7': PortAllUser7,

                    # alluser
                                                    'UnloadReidlinAllUser': NanCheck(Reid_infoAll[0][0]),
                                                    'UnloadReidtrampAllUser': NanCheck(Reid_infoAll[0][1]),
                                                    'UnloadReidSumAllUser': (NanCheck(Reid_infoAll[0][0]) + NanCheck(Reid_infoAll[0][1])),
                                                    'LoadingReidLinAllUser': NanCheck(Reid_infoAll[0][2]),
                                                    'LoadingReidTrampAllUser': NanCheck(Reid_infoAll[0][3]),
                                                    'LoadingReidSumAllUser': (NanCheck(Reid_infoAll[0][2]) + NanCheck(Reid_infoAll[0][3])),
                                                    'ReidAllAll' : ReidAllAll,
                                                    'UnloadPortlinAllUser': NanCheck(Reid_infoAll[0][4]),
                                                    'UnloadPorttrampAllUser': NanCheck(Reid_infoAll[0][5]),
                                                    'UnloadPortSumAllUser': (NanCheck(Reid_infoAll[0][4]) + NanCheck(Reid_infoAll[0][5])),
                                                    'LoadingPortLinAllUser': NanCheck(Reid_infoAll[0][6]),
                                                    'LoadingPortTrampAllUser': NanCheck(Reid_infoAll[0][7]),
                                                    'LoadingPortSumAllUser': (NanCheck(Reid_infoAll[0][6]) + NanCheck(Reid_infoAll[0][7])),
                                                    'PortAllAll': PortAllAll,

                    #6 table (2 in dataset)

                                                    'MaxWarehouseQty8User': MaxWarehouseQty8User,
                                                    'MaxWarehouseQty9User': MaxWarehouseQty9User,
                                                    'MaxWarehouseQty10User': MaxWarehouseQty10User,
                                                    'MaxWarehouseQty11User': MaxWarehouseQty11User,
                                                    'MaxWarehouseQty12User': MaxWarehouseQty12User,
                                                    'MaxWarehouseQty13User': MaxWarehouseQty13User,
                                                    'MaxWarehouseQty14User': MaxWarehouseQty14User,
                                                    'MaxWarehouseQtyAllNotST': MaxWarehouseQtyAllNotST,

                                                    'FactLoad8User' : NanCheck(TransportUserInfo8[0][8]),
                                                    'FactLoad9User': NanCheck(TransportUserInfo9[0][8]),
                                                    'FactLoad10User': NanCheck(TransportUserInfo10[0][8]),
                                                    'FactLoad11User': NanCheck(TransportUserInfo11[0][8]),
                                                    'FactLoad12User': NanCheck(TransportUserInfo12[0][8]),
                                                    'FactLoad13User': NanCheck(TransportUserInfo13[0][8]),
                                                    'FactLoad14User': NanCheck(TransportUserInfo14[0][8]),
                                                    'FactLoadAllNotST': NanCheck(TransportUserInfoNotSTAll[0][8]),


                                                    'Percent8User': TransportUserInfo8Percent,
                                                    'Percent9User': TransportUserInfo9Percent,
                                                    'Percent10User': TransportUserInfo10Percent,
                                                    'Percent11User': TransportUserInfo11Percent,
                                                    'Percent12User': TransportUserInfo12Percent,
                                                    'Percent13User': TransportUserInfo13Percent,
                                                    'Percent14User': TransportUserInfo14Percent,
                                                    'PercentAllNotST': TransportUserInfoAllPercent,


                                                    'TransportQty8UserIn': NanCheck(TransportUserInfo8[0][4]) +
                                                                           NanCheck(TransportUserInfo8[0][5]) +
                                                                           NanCheck(TransportUserInfo8[0][6]),
                                                    'TransportQty9UserIn': NanCheck(TransportUserInfo9[0][4]) +
                                                                           NanCheck(TransportUserInfo9[0][5]) +
                                                                           NanCheck(TransportUserInfo9[0][6]),
                                                    'TransportQty10UserIn': NanCheck(TransportUserInfo10[0][4]) +
                                                                            NanCheck(TransportUserInfo10[0][5]) +
                                                                            NanCheck(TransportUserInfo10[0][6]),
                                                    'TransportQty11UserIn': NanCheck(TransportUserInfo11[0][4]) +
                                                                            NanCheck(TransportUserInfo11[0][5]) +
                                                                            NanCheck(TransportUserInfo11[0][6]),
                                                    'TransportQty12UserIn': NanCheck(TransportUserInfo12[0][4]) +
                                                                            NanCheck(TransportUserInfo12[0][5]) +
                                                                            NanCheck(TransportUserInfo12[0][6]),
                                                    'TransportQty13UserIn': NanCheck(TransportUserInfo13[0][4]) +
                                                                            NanCheck(TransportUserInfo13[0][5]) +
                                                                            NanCheck(TransportUserInfo13[0][6]),
                                                    'TransportQty14UserIn': NanCheck(TransportUserInfo14[0][4]) +
                                                                            NanCheck(TransportUserInfo14[0][5]) +
                                                                            NanCheck(TransportUserInfo14[0][6]),
                                                    'TransportQtyAllIn': NanCheck(TransportUserInfoNotSTAll[0][4]) +
                                                                         NanCheck(TransportUserInfoNotSTAll[0][5]) +
                                                                         NanCheck(TransportUserInfoNotSTAll[0][6]),


                                                    'TransportQty8UserOut': NanCheck(TransportUserInfo8[0][0]) +
                                                                            NanCheck(TransportUserInfo8[0][1]) +
                                                                            NanCheck(TransportUserInfo8[0][2]),
                                                    'TransportQty9UserOut': NanCheck(TransportUserInfo9[0][0]) +
                                                                            NanCheck(TransportUserInfo9[0][1]) +
                                                                            NanCheck(TransportUserInfo9[0][2]),
                                                    'TransportQty10UserOut': NanCheck(TransportUserInfo10[0][0]) +
                                                                             NanCheck(TransportUserInfo10[0][1]) +
                                                                             NanCheck(TransportUserInfo10[0][2]),
                                                    'TransportQty11UserOut': NanCheck(TransportUserInfo11[0][0]) +
                                                                             NanCheck(TransportUserInfo11[0][1]) +
                                                                             NanCheck(TransportUserInfo11[0][2]),
                                                    'TransportQty12UserOut': NanCheck(TransportUserInfo12[0][0]) +
                                                                             NanCheck(TransportUserInfo12[0][1]) +
                                                                             NanCheck(TransportUserInfo12[0][2]),
                                                    'TransportQty13UserOut': NanCheck(TransportUserInfo13[0][0]) +
                                                                             NanCheck(TransportUserInfo13[0][1]) +
                                                                             NanCheck(TransportUserInfo13[0][2]),
                                                    'TransportQty14UserOut': NanCheck(TransportUserInfo14[0][0]) +
                                                                             NanCheck(TransportUserInfo14[0][1]) +
                                                                             NanCheck(TransportUserInfo14[0][2]),
                                                    'TransportQtyAllOut': NanCheck(TransportUserInfoNotSTAll[0][0]) +
                                                                          NanCheck(TransportUserInfoNotSTAll[0][1]) +
                                                                          NanCheck(TransportUserInfoNotSTAll[0][2]),

                                                    'TransportCalculated8': TransportCalculated8,
                                                    'TransportCalculated9': TransportCalculated9,
                                                    'TransportCalculated10': TransportCalculated10,
                                                    'TransportCalculated11': TransportCalculated11,
                                                    'TransportCalculated12': TransportCalculated12,
                                                    'TransportCalculated13': TransportCalculated13,
                                                    'TransportCalculated14': TransportCalculated14,
                                                    'TransportCalculatedAllNotST': TransportCalculatedAllNotST,

                                                    'user': request.user})
    else:
        return redirect('home')

# datepick for use in dataset
@login_required(login_url='')
def datepick_admin(request):
    if request.user.id == 1:
        if request.method == 'POST':
            date1 = request.POST['date1']
            request.session['parameters'] = {'date1': ((dt.datetime.strptime(date1, '%Y-%m-%d') - DAYDELTA).strftime('%Y-%m-%d'))}
            return redirect(dataset)
        return render(request, 'datepick_admin.html' , {'currentdate':(dt.datetime.now()).strftime('%Y-%m-%d')})
    else:
        return redirect('home')


@login_required(login_url='')
def table1_upload(request):
    if request.method == 'POST':
        # file = request.FILES.get('file')
        # if file is not None:
        #     if file.name.endswith('.xlsx'):
        #         fs = FileSystemStorage()
        #         filename = fs.save(file.name, file)
        #         try:
        #             # Чтение xlsx файла
        #             df = pd.read_excel(fs.path(filename))
        #             # Игнорирование первых трёх строк
        #             df = df.iloc[2:]
        #             # Преобразование DataFrame в списки столбцов (В дальнейшем для новых столбцов таблицы создавать новые поля в таблице БД, а также добавлять ниже в переменные)
        #             ImportIn = df.iloc[:, 0].tolist()  # в т.ч. импорт Прибыло
        #             ImportOut = df.iloc[:, 1].tolist()  # в т.ч. импорт Убыло
        #             ExportIn = df.iloc[:, 2].tolist()  # в т.ч. экспорт Прибыло
        #             ExportOut = df.iloc[:, 3].tolist()  # в т.ч. экспорт Убыло
        #             TransitIn = df.iloc[:, 4].tolist()  # в т.ч. транзит Прибыло
        #             TransitOut = df.iloc[:, 5].tolist()  # в т.ч. транзит Убыло
        #             ExportEmpty = df.iloc[:, 6].tolist()  # в т.ч. экспорт порожние
        #             OtherEmpty = df.iloc[:, 7].tolist()  # в т.ч. прочие порожние
        #             UnloadReid = df.iloc[:, 8].tolist()  # На рейде в ожидании Выгрузки
        #             LoadingReid = df.iloc[:, 9].tolist()  # На рейде в ожидании Погрузки
        #             UnloadPort = df.iloc[:, 10].tolist()  # На подходах к порту для Выгрузки
        #             LoadingPort = df.iloc[:, 11].tolist()  # На подходах к порту для Погрузки
        #             fs.delete(filename)
        #
        #             # Дальнейшая обработка данных
        #             request.session['parameters'] = {
        #                 'ImportIn': ImportIn,
        #                 'ImportOut': ImportOut,
        #                 'ExportIn': ExportIn,
        #                 'ExportOut': ExportOut,
        #                 'TransitIn': TransitIn,
        #                 'TransitOut': TransitOut,
        #                 'ExportEmpty': ExportEmpty,
        #                 'OtherEmpty': OtherEmpty,
        #                 'UnloadReid': UnloadReid,
        #                 'LoadingReid': LoadingReid,
        #                 'UnloadPort': UnloadPort,
        #                 'LoadingPort': LoadingPort
        #             }
        #             # Сохраните сессию, чтобы сгенерировать сессионный ключ
        #             request.session.save()
        #
        #             # Получить текущий session ID
        #             session_id = request.session.session_key
        #             # Создайте URL-адрес перенаправления с этим session ID
        #             redirect_url = f'/table1_data/?session_id={session_id}'
        #             return redirect(redirect_url)
        #         except:
        #             fs.delete(filename)
        #             return render(request, 'error.html', {'ErrorText': 'Ошибка выгрузки данных'})
        #     else:
        #         return render(request, 'error.html', {'ErrorText': 'Неверный формат файла'})
        # else:
        return redirect('table1_data')
    return render(request, 'table1_upload.html')

@login_required(login_url='')
def table2_upload(request):
    if request.method == 'POST':
        return redirect('table2_data')
    return render(request, 'table2_upload.html')

@login_required(login_url='')
def table2_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        #WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        ContainerTrain = params.get('ContainerTrain')
        ContainerAuto = params.get('ContainerAuto')
        ContainerAutoQty = params.get('ContainerAutoQty')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            ContainerTrain = [request.POST['ContainerTrain']]
            ContainerAuto = [request.POST['ContainerAuto']]
            ContainerAutoQty = [request.POST['ContainerAutoQty']]
            request.session['parameters'] = {
                'date2' : date2,
                'ContainerTrain': ContainerTrain,
                'ContainerAuto': ContainerAuto,
                'ContainerAutoQty': ContainerAutoQty
            }
            redirect_url = f'/success_table2/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table2_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'ContainerTrain': ContainerTrain[0],
                                                        'ContainerAuto': ContainerAuto[0],
                                                        'ContainerAutoQty': ContainerAutoQty[0],
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            ContainerTrain = [request.POST['ContainerTrain']]
            ContainerAuto = [request.POST['ContainerAuto']]
            ContainerAutoQty = [request.POST['ContainerAutoQty']]
            request.session['parameters'] = {
                'date2': date2,
                'ContainerTrain': ContainerTrain,
                'ContainerAuto': ContainerAuto,
                'ContainerAutoQty': ContainerAutoQty,
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table2/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            conn = connection()
            Userdata = getContaunerUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table2_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ContainerTrain': NanCheck(Userdata[0][0]),
                'ContainerAuto': NanCheck(Userdata[0][1]),
                'ContainerAutoQty': NanCheck(Userdata[0][2]),
            })
        except:
            return render(request, 'table2_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ContainerTrain': 0,
                'ContainerAuto': 0,
                'ContainerAutoQty': 0
            })


@transaction.atomic
@login_required(login_url='')
def success_table2(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        ContainerTrain = params.get('ContainerTrain')
        ContainerAuto = params.get('ContainerAuto')
        ContainerAutoQty = params.get('ContainerAutoQty')
        ContainerTrain[0] = NanCheck(ContainerTrain[0])
        ContainerAuto[0] = NanCheck(ContainerAuto[0])
        ContainerAutoQty[0] = NanCheck(ContainerAutoQty[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = (DailyMonitoringUserContainers.objects.filter(
            date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                db_userid = request.user.id)
        .update(
                db_container_train=int(ContainerTrain[0]),
                db_container_auto=int(ContainerAuto[0]),
                db_container_auto_qty=int(ContainerAutoQty[0])
        ))
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserContainers.objects.create(
                                        date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                        db_userid = request.user.id,
                                        db_container_train= int(ContainerTrain[0]),
                                        db_container_auto = int(ContainerAuto[0]),
                                        db_container_auto_qty = int(ContainerAutoQty[0])
                                        )

        return render(request, 'success_table2.html', {
                                                'date2': date2[0][0],
                                                'ContainerTrain': ContainerTrain[0],
                                                'ContainerAuto': ContainerAuto[0],
                                                'ContainerAutoQty': ContainerAutoQty[0],
                                                'user': request.user})


@login_required(login_url='')
def table3_upload(request):
    if request.method == 'POST':
        return redirect('table3_data')
    return render(request, 'table3_upload.html')

@login_required(login_url='')
def table3_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        Wagons = params.get('Wagons')
        WagonsOut = params.get('WagonsOut')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            Wagons = [request.POST['Wagons']]
            WagonsOut = [request.POST['WagonsOut']]
            request.session['parameters'] = {
                'date2' : date2,
                'Wagons': Wagons,
                'WagonsOut': WagonsOut,
            }
            redirect_url = f'/success_table3/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table3_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'Wagons': Wagons[0],
                                                        'WagonsOut': WagonsOut[0],
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            Wagons = [request.POST['Wagons']]
            WagonsOut = [request.POST['WagonsOut']]
            request.session['parameters'] = {
                'date2': date2,
                'Wagons': Wagons,
                'WagonsOut': WagonsOut,
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table3/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            conn = connection()
            Userdata = getWagonsUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table3_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'Wagons': NanCheck(Userdata[0][0]),
                'WagonsOut': NanCheck(Userdata[0][1]),
            })
        except:
            return render(request, 'table3_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'Wagons': 0,
                'WagonsOut': 0,
            })


@transaction.atomic
@login_required(login_url='')
def success_table3(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        Wagons = params.get('Wagons')
        WagonsOut = params.get('WagonsOut')
        Wagons[0] = NanCheck(Wagons[0])
        WagonsOut[0] = NanCheck(WagonsOut[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserWagons.objects.filter(
                date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                db_userid = request.user.id).update(
                    db_wagons=int(Wagons[0]),
                    db_wagons_out=int(WagonsOut[0]),
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserWagons.objects.create(
                                        date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                        db_userid = request.user.id,
                                        db_wagons= int(Wagons[0]),
                                        db_wagons_out = int(WagonsOut[0]),
                                        )

        return render(request, 'success_table3.html', {
                                                'date2': date2[0][0],
                                                'Wagons': Wagons[0],
                                                'WagonsOut': WagonsOut[0],
                                                'user': request.user})



@login_required(login_url='')
def table4_upload(request):
    if request.method == 'POST':
        return redirect('table4_data')
    return render(request, 'table4_upload.html')

@login_required(login_url='')
def table4_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        Wagons_FE = params.get('Wagons_FE')
        WagonsOut_FE = params.get('WagonsOut_FE')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            Wagons_FE = [request.POST['Wagons_FE']]
            WagonsOut_FE = [request.POST['WagonsOut_FE']]
            request.session['parameters'] = {
                'date2' : date2,
                'Wagons_FE': Wagons_FE,
                'WagonsOut_FE': WagonsOut_FE,
            }
            redirect_url = f'/success_table4/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table4_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'Wagons_FE': Wagons_FE[0],
                                                        'WagonsOut_FE': WagonsOut_FE[0],
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            Wagons_FE = [request.POST['Wagons_FE']]
            WagonsOut_FE = [request.POST['WagonsOut_FE']]
            request.session['parameters'] = {
                'date2': date2,
                'Wagons_FE': Wagons_FE,
                'WagonsOut_FE': WagonsOut_FE,
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table4/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            conn = connection()
            Userdata = getWagonsUserInfoFromDBFE(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table4_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'Wagons_FE': NanCheck(Userdata[0][0]),
                'WagonsOut_FE': NanCheck(Userdata[0][1]),
            })
        except:
            return render(request, 'table4_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'Wagons_FE': 0,
                'WagonsOut_FE': 0,
            })


@transaction.atomic
@login_required(login_url='')
def success_table4(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        Wagons_FE = params.get('Wagons_FE')
        WagonsOut_FE = params.get('WagonsOut_FE')
        Wagons_FE[0] = NanCheck(Wagons_FE[0])
        WagonsOut_FE[0] = NanCheck(WagonsOut_FE[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserWagonsFE.objects.filter(
            date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
            db_userid = request.user.id).update(
                db_wagons_fe=int(Wagons_FE[0]),
                db_wagons_out_fe=int(WagonsOut_FE[0]),
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserWagonsFE.objects.create(
                                          date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                          db_userid = request.user.id,
                                          db_wagons_fe= int(Wagons_FE[0]),
                                          db_wagons_out_fe = int(WagonsOut_FE[0]),
                                          )

        return render(request, 'success_table4.html', {
                                                'date2': date2[0][0],
                                                'Wagons_FE': Wagons_FE[0],
                                                'WagonsOut_FE': WagonsOut_FE[0],
                                                'user': request.user})


@login_required(login_url='')
def table5_upload(request):
    if request.method == 'POST':
        return redirect('table5_data')
    return render(request, 'table5_upload.html')

@login_required(login_url='')
def table5_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        FittingPlatformOut = params.get('FittingPlatformOut')
        SemiwagonOut = params.get('SemiwagonOut')
        AutoOut = params.get('AutoOut')
        SeaOut = params.get('SeaOut')
        FittingPlatformIn = params.get('FittingPlatformIn')
        SemiwagonIn = params.get('SemiwagonIn')
        AutoIn = params.get('AutoIn')
        SeaIn = params.get('SeaIn')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            FittingPlatformOut = [request.POST['FittingPlatformOut']]
            SemiwagonOut = [request.POST['SemiwagonOut']]
            AutoOut = [request.POST['AutoOut']]
            SeaOut = [request.POST['SeaOut']]
            FittingPlatformIn = [request.POST['FittingPlatformIn']]
            SemiwagonIn = [request.POST['SemiwagonIn']]
            AutoIn = [request.POST['AutoIn']]
            SeaIn = [request.POST['SeaIn']]
            request.session['parameters'] = {
                'date2' : date2,
                'FittingPlatformOut': FittingPlatformOut,
                'SemiwagonOut': SemiwagonOut,
                'AutoOut': AutoOut,
                'SeaOut': SeaOut,
                'FittingPlatformIn': FittingPlatformIn,
                'SemiwagonIn': SemiwagonIn,
                'AutoIn': AutoIn,
                'SeaIn': SeaIn,
            }
            redirect_url = f'/success_table5/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table5_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'FittingPlatformOut': FittingPlatformOut[0],
                                                        'SemiwagonOut': SemiwagonOut[0],
                                                        'AutoOut': AutoOut[0],
                                                        'SeaOut': SeaOut[0],
                                                        'FittingPlatformIn': FittingPlatformIn[0],
                                                        'SemiwagonIn': SemiwagonIn[0],
                                                        'AutoIn': AutoIn[0],
                                                        'SeaIn': SeaIn[0],
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            FittingPlatformOut = [request.POST['FittingPlatformOut']]
            SemiwagonOut = [request.POST['SemiwagonOut']]
            AutoOut = [request.POST['AutoOut']]
            SeaOut = [request.POST['SeaOut']]
            FittingPlatformIn = [request.POST['FittingPlatformIn']]
            SemiwagonIn = [request.POST['SemiwagonIn']]
            AutoIn = [request.POST['AutoIn']]
            SeaIn = [request.POST['SeaIn']]
            request.session['parameters'] = {
                'date2': date2,
                'FittingPlatformOut': FittingPlatformOut,
                'SemiwagonOut': SemiwagonOut,
                'AutoOut': AutoOut,
                'SeaOut': SeaOut,
                'FittingPlatformIn': FittingPlatformIn,
                'SemiwagonIn': SemiwagonIn,
                'AutoIn': AutoIn,
                'SeaIn': SeaIn,
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table5/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            conn = connection()
            Userdata = getTransportUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table5_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'FittingPlatformOut': NanCheck(Userdata[0][0]),
                'SemiwagonOut': NanCheck(Userdata[0][1]),
                'AutoOut': NanCheck(Userdata[0][2]),
                'SeaOut': NanCheck(Userdata[0][3]),
                'FittingPlatformIn': NanCheck(Userdata[0][4]),
                'SemiwagonIn': NanCheck(Userdata[0][5]),
                'AutoIn': NanCheck(Userdata[0][6]),
                'SeaIn': NanCheck(Userdata[0][7]),
            })
        except:
            return render(request, 'table5_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'FittingPlatformOut': 0,
                'SemiwagonOut': 0,
                'AutoOut': 0,
                'SeaOut': 0,
                'FittingPlatformIn': 0,
                'SemiwagonIn': 0,
                'AutoIn': 0,
                'SeaIn': 0,
            })


@transaction.atomic
@login_required(login_url='')
def success_table5(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        FittingPlatformOut = params.get('FittingPlatformOut')
        SemiwagonOut = params.get('SemiwagonOut')
        AutoOut = params.get('AutoOut')
        SeaOut = params.get('SeaOut')
        FittingPlatformIn = params.get('FittingPlatformIn')
        SemiwagonIn = params.get('SemiwagonIn')
        AutoIn = params.get('AutoIn')
        SeaIn = params.get('SeaIn')
        FittingPlatformOut[0] = NanCheck(FittingPlatformOut[0])
        SemiwagonOut[0] = NanCheck(SemiwagonOut[0])
        AutoOut[0] = NanCheck(AutoOut[0])
        SeaOut[0] = NanCheck(SeaOut[0])
        FittingPlatformIn[0] = NanCheck(FittingPlatformIn[0])
        SemiwagonIn[0] = NanCheck(SemiwagonIn[0])
        AutoIn[0] = NanCheck(AutoIn[0])
        SeaIn[0] = NanCheck(SeaIn[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserTransport.objects.filter(
            date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
            db_userid = request.user.id).update(
                db_fittingplatform_out=int(FittingPlatformOut[0]),
                db_semiwagon_out=int(SemiwagonOut[0]),
                db_auto_out=int(AutoOut[0]),
                db_sea_out=int(SeaOut[0]),
                db_fittingplatform_in=int(FittingPlatformIn[0]),
                db_semiwagon_in=int(SemiwagonIn[0]),
                db_auto_in=int(AutoIn[0]),
                db_sea_in=int(SeaIn[0]),
                db_factload=0,
                db_reload=0,
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserTransport.objects.create(
                                        date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                        db_userid = request.user.id,
                                        db_fittingplatform_out=int(FittingPlatformOut[0]),
                                        db_semiwagon_out=int(SemiwagonOut[0]),
                                        db_auto_out=int(AutoOut[0]),
                                        db_sea_out=int(SeaOut[0]),
                                        db_fittingplatform_in=int(FittingPlatformIn[0]),
                                        db_semiwagon_in=int(SemiwagonIn[0]),
                                        db_auto_in=int(AutoIn[0]),
                                        db_sea_in=int(SeaIn[0]),
                                        db_factload=0,
                                        db_reload=0,
                                        )

        return render(request, 'success_table5.html', {
                                                'date2': date2[0][0],
                                                'FittingPlatformOut': FittingPlatformOut[0],
                                                'SemiwagonOut': SemiwagonOut[0],
                                                'AutoOut': AutoOut[0],
                                                'SeaOut': SeaOut[0],
                                                'FittingPlatformIn': FittingPlatformIn[0],
                                                'SemiwagonIn': SemiwagonIn[0],
                                                'AutoIn': AutoIn[0],
                                                'SeaIn': SeaIn[0],
                                                'user': request.user})


@login_required(login_url='')
def table6_upload(request):
    if request.method == 'POST':
        return redirect('table6_data')
    return render(request, 'table6_upload.html')

@login_required(login_url='')
def table6_data(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # WORK IN PROGRESS
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        FittingPlatformOut = params.get('FittingPlatformOut')
        SemiwagonOut = params.get('SemiwagonOut')
        AutoOut = params.get('AutoOut')
        FittingPlatformIn = params.get('FittingPlatformIn')
        SemiwagonIn = params.get('SemiwagonIn')
        AutoIn = params.get('AutoIn')
        FactLoad = params.get('FactLoad')
        Reload = params.get('Reload')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            FittingPlatformOut = [request.POST['FittingPlatformOut']]
            SemiwagonOut = [request.POST['SemiwagonOut']]
            AutoOut = [request.POST['AutoOut']]
            FittingPlatformIn = [request.POST['FittingPlatformIn']]
            SemiwagonIn = [request.POST['SemiwagonIn']]
            AutoIn = [request.POST['AutoIn']]
            FactLoad = [request.POST['FactLoad']]
            Reload = [request.POST['Reload']]
            request.session['parameters'] = {
                'date2' : date2,
                'FittingPlatformOut': FittingPlatformOut,
                'SemiwagonOut': SemiwagonOut,
                'AutoOut': AutoOut,
                'FittingPlatformIn': FittingPlatformIn,
                'SemiwagonIn': SemiwagonIn,
                'AutoIn': AutoIn,
                'FactLoad': FactLoad,
                'Reload': Reload,
            }
            redirect_url = f'/success_table5/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'table6_data.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'FittingPlatformOut': FittingPlatformOut[0],
                                                        'SemiwagonOut': SemiwagonOut[0],
                                                        'AutoOut': AutoOut[0],
                                                        'FittingPlatformIn': FittingPlatformIn[0],
                                                        'SemiwagonIn': SemiwagonIn[0],
                                                        'AutoIn': AutoIn[0],
                                                        'FactLoad': FactLoad[0],
                                                        'Reload': Reload[0],
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            FittingPlatformOut = [request.POST['FittingPlatformOut']]
            SemiwagonOut = [request.POST['SemiwagonOut']]
            AutoOut = [request.POST['AutoOut']]
            FittingPlatformIn = [request.POST['FittingPlatformIn']]
            SemiwagonIn = [request.POST['SemiwagonIn']]
            AutoIn = [request.POST['AutoIn']]
            FactLoad = [request.POST['FactLoad']]
            Reload = [request.POST['Reload']]
            request.session['parameters'] = {
                'date2': date2,
                'FittingPlatformOut': FittingPlatformOut,
                'SemiwagonOut': SemiwagonOut,
                'AutoOut': AutoOut,
                'FittingPlatformIn': FittingPlatformIn,
                'SemiwagonIn': SemiwagonIn,
                'AutoIn': AutoIn,
                'FactLoad': FactLoad,
                'Reload': Reload,
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success_table6/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            #WORK IN PROGRESS
            conn = connection()
            Userdata = getUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'),conn)
            conn.close()
            return render(request, 'table6_data.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'FittingPlatformOut': (Userdata[0][0]),
                'SemiwagonOut': NanCheck(Userdata[0][1]),
                'AutoOut': NanCheck(Userdata[0][2]),
                'FittingPlatformIn': NanCheck(Userdata[0][4]),
                'SemiwagonIn': NanCheck(Userdata[0][5]),
                'AutoIn': NanCheck(Userdata[0][6]),
                'FactLoad': NanCheck(Userdata[0][8]),
                'Reload': NanCheck(Userdata[0][9]),
            })
        except:
            return render(request, 'table6_data.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'FittingPlatformOut': 0,
                'SemiwagonOut': 0,
                'AutoOut': 0,
                'FittingPlatformIn': 0,
                'SemiwagonIn': 0,
                'AutoIn': 0,
                'FactLoad': 0,
                'Reload': 0,
            })


@transaction.atomic
@login_required(login_url='')
def success_table6(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        FittingPlatformOut = params.get('FittingPlatformOut')
        SemiwagonOut = params.get('SemiwagonOut')
        AutoOut = params.get('AutoOut')
        FittingPlatformIn = params.get('FittingPlatformIn')
        SemiwagonIn = params.get('SemiwagonIn')
        AutoIn = params.get('AutoIn')
        FactLoad = params.get('FactLoad')
        Reload = params.get('Reload')
        FittingPlatformOut[0] = NanCheck(FittingPlatformOut[0])
        SemiwagonOut[0] = NanCheck(SemiwagonOut[0])
        AutoOut[0] = NanCheck(AutoOut[0])
        FittingPlatformIn[0] = NanCheck(FittingPlatformIn[0])
        SemiwagonIn[0] = NanCheck(SemiwagonIn[0])
        AutoIn[0] = NanCheck(AutoIn[0])
        FactLoad[0] = NanCheck(FactLoad[0])
        Reload[0] = NanCheck(Reload[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserTransport.objects.filter(
            date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
            db_userid = request.user.id).update(
                db_fittingplatform_out=int(FittingPlatformOut[0]),
                db_semiwagon_out=int(SemiwagonOut[0]),
                db_auto_out=int(AutoOut[0]),
                db_sea_out=0,
                db_fittingplatform_in=int(FittingPlatformIn[0]),
                db_semiwagon_in=int(SemiwagonIn[0]),
                db_auto_in=int(AutoIn[0]),
                db_sea_in=0,
                db_factload=int(FactLoad[0]),
                db_reload=int(Reload[0])
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserTransport.objects.create(
                date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                        db_userid = request.user.id,
                                        db_fittingplatform_out=int(FittingPlatformOut[0]),
                                        db_semiwagon_out=int(SemiwagonOut[0]),
                                        db_auto_out=int(AutoOut[0]),
                                        db_sea_out=0,
                                        db_fittingplatform_in=int(FittingPlatformIn[0]),
                                        db_semiwagon_in=int(SemiwagonIn[0]),
                                        db_auto_in=int(AutoIn[0]),
                                        db_sea_in=0,
                                        db_factload=int(FactLoad[0]),
                                        db_reload=int(Reload[0])
                                        )

        return render(request, 'success_table6.html', {
                                                'date2': date2[0][0],
                                                'FittingPlatformOut': FittingPlatformOut[0],
                                                'SemiwagonOut': SemiwagonOut[0],
                                                'AutoOut': AutoOut[0],
                                                'FittingPlatformIn': FittingPlatformIn[0],
                                                'SemiwagonIn': SemiwagonIn[0],
                                                'AutoIn': AutoIn[0],
                                                'FactLoad': FactLoad[0],
                                                'Reload': Reload[0],
                                                'user': request.user})